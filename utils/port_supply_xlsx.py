"""utils/port_supply_xlsx.py — single .xlsx workbook bundling all 5 views.

Where ``utils.port_supply_csv`` produces 5 separate CSV files, this
module produces a single Excel workbook with 5 sheets — easier to share
+ navigate for analysts who want cross-sheet pivots without juggling
multiple files.

Sheets (one per view, named identically to the CLI's view names):

  * ``summary``    — per-port summary + analytics (deficit rank,
                     regional context, exposure concentration HHI)
  * ``exposure``   — flat port × company join + share + rank within port
  * ``footprint``  — reverse axis (ticker × port) + deficit share +
                     concentration HHI + top region
  * ``regional``   — one row per region with aggregate stats
  * ``watchlist``  — deficit-only subset with action + why columns

Plus an ``overview`` sheet at position 0 that summarises the snapshot:
container type, timestamp, sheet inventory + per-sheet row counts.

Every sheet:
  * Bold header row + frozen so it stays visible while scrolling
  * Column widths auto-sized from header + a sample of body cells
  * Numeric columns formatted (4 decimal places for weights, 2 for
    deficit days, 1 for utilization %)

Pure function — returns the workbook as bytes (no file I/O), so the
UI download button + the CLI both consume it through the same path.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Iterable


__all__ = [
    "WORKBOOK_SHEETS",
    "build_workbook",
]


# Canonical sheet order. Mirrors the CSV view-name set so users can
# round-trip between CLI exports and the workbook without remapping.
WORKBOOK_SHEETS: tuple[str, ...] = (
    "overview",
    "summary",
    "exposure",
    "footprint",
    "regional",
    "watchlist",
)


# ---------------------------------------------------------------------------
# Internal: parse a CSV string back into header + rows, stripping our
# UTF-8 BOM + comment-header lines. Lets us reuse the existing CSV
# exporters as the single source of truth for column layout instead of
# duplicating per-view writers.
# ---------------------------------------------------------------------------


def _parse_csv_payload(csv_text: str) -> tuple[list[str], list[list[str]]]:
    """Return (header_row, data_rows) from a CSV string emitted by
    ``utils.port_supply_csv``. Strips BOM + comment lines."""
    if csv_text.startswith("﻿"):
        csv_text = csv_text[1:]
    lines = [
        line for line in csv_text.split("\n")
        if line and not line.startswith("# ")
    ]
    if not lines:
        return [], []
    reader = csv.reader(io.StringIO("\n".join(lines)))
    rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


# ---------------------------------------------------------------------------
# Column-width estimator + numeric coercion
# ---------------------------------------------------------------------------


def _autosize_column(header: str, sample_cells: Iterable[str]) -> int:
    """Pick a column width = max(header, max(sample cell)) + a small
    cushion. Caps at 60 so a long summary cell doesn't blow out the
    layout."""
    widths = [len(header)]
    for s in sample_cells:
        if s is None:
            continue
        widths.append(len(str(s)))
    return min(60, max(widths) + 2)


# Per-column number-format hint registry. Keyed by header substring so
# the same hint applies across all five views (e.g. any column whose
# name contains "deficit_days" gets the days format).
_NUMBER_FORMATS: list[tuple[str, str]] = [
    ("deficit_days",   "+0.00;-0.00"),       # signed days
    ("utilization",    "0.0"),                # percentage-ish
    ("exposure_weight", "0.000000"),          # 6 decimal places
    ("total_exposure", "0.000000"),
    ("deficit_weighted_score", "0.000000"),
    ("hhi",            "0.0000"),
    ("share",          "0.0000"),
    ("lat",            "0.00"),
    ("lon",            "0.00"),
]


def _format_hint_for(header: str) -> str:
    """Return the openpyxl number_format string for a header, or '' if
    no hint applies (cell stays as General format)."""
    lower = header.lower()
    for substr, fmt in _NUMBER_FORMATS:
        if substr in lower:
            return fmt
    return ""


def _coerce_numeric(value: str) -> object:
    """Convert string cells that look numeric to floats / ints so
    openpyxl writes them as numbers (sortable, formula-friendly) rather
    than text. Falls back to the original string for anything else."""
    if value is None or value == "":
        return ""
    stripped = value.strip()
    # Handle the signed-deficit format "+5.20" / "-3.00"
    cleaned = stripped
    try:
        if "." in cleaned:
            return float(cleaned)
        return int(cleaned)
    except ValueError:
        return value


# ---------------------------------------------------------------------------
# Sheet writer — one per CSV string
# ---------------------------------------------------------------------------


def _write_csv_to_sheet(ws, csv_text: str) -> int:
    """Parse the CSV string, lay it out into ``ws`` with bold header +
    frozen row + autosized columns + numeric coercion + number-format
    hints. Returns the body row count (excludes header)."""
    from openpyxl.styles import Font

    header, rows = _parse_csv_payload(csv_text)
    if not header:
        return 0

    # Header row
    bold = Font(bold=True)
    for col_idx, name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = bold
    ws.freeze_panes = "A2"

    # Data rows — coerce numerics, apply per-column format hints.
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            if col_idx > len(header):
                continue
            coerced = _coerce_numeric(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=coerced)
            hint = _format_hint_for(header[col_idx - 1])
            if hint and isinstance(coerced, (int, float)):
                cell.number_format = hint

    # Column widths — sample up to 30 body cells per column.
    from openpyxl.utils import get_column_letter
    for col_idx, name in enumerate(header, 1):
        sample = [
            row[col_idx - 1] if col_idx - 1 < len(row) else ""
            for row in rows[:30]
        ]
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            _autosize_column(name, sample)
        )

    return len(rows)


# ---------------------------------------------------------------------------
# Overview sheet — snapshot metadata + sheet inventory + row counts
# ---------------------------------------------------------------------------


def _write_overview_sheet(
    ws,
    *,
    container_type: str,
    threshold_days: float,
    sheet_rows: dict[str, int],
) -> None:
    from openpyxl.styles import Font
    bold = Font(bold=True)

    rows: list[tuple[str, object]] = [
        ("Port Supply Lines — workbook snapshot", ""),
        ("", ""),
        ("Generated (UTC)",
         datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")),
        ("Container type",  container_type),
        ("Watchlist threshold (days)", float(threshold_days)),
        ("Source", "processing.port_supply_lines (modeled)"),
        ("Schema version", 2),
        ("", ""),
        ("Sheet", "Rows"),
    ]
    for label, value in rows:
        c = ws.cell(row=ws.max_row + 1, column=1, value=label)
        c.font = bold
        ws.cell(row=ws.max_row, column=2, value=value)

    # Sheet inventory + row counts (data rows, header excluded).
    for sheet_name in WORKBOOK_SHEETS:
        if sheet_name == "overview":
            continue
        n = sheet_rows.get(sheet_name, 0)
        ws.cell(row=ws.max_row + 1, column=1, value=sheet_name)
        ws.cell(row=ws.max_row, column=2, value=n)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 36


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def build_workbook(
    chains: Iterable,
    footprints: Iterable,
    *,
    container_type: str = "40FT_DRY",
    threshold_days: float = -3.0,
) -> bytes:
    """Build the 6-sheet workbook + return it as bytes.

    Parameters
    ----------
    chains:
        Output of ``processing.port_supply_lines.build_port_supply_chains``.
    footprints:
        Output of ``processing.port_supply_lines.build_company_port_footprints``.
    container_type:
        Surfaced in the overview sheet + propagates into per-CSV
        metadata headers.
    threshold_days:
        Watchlist firing threshold; surfaced in the overview sheet and
        used by the watchlist CSV.

    Returns
    -------
    bytes
        The serialized .xlsx workbook. Caller writes to disk / pipes
        into ``st.download_button`` / sends over HTTP.
    """
    from openpyxl import Workbook
    from utils.port_supply_csv import (
        chains_to_deficit_watchlist_csv,
        chains_to_exposure_csv,
        chains_to_regional_rollup_csv,
        chains_to_summary_csv,
        footprints_to_csv,
    )

    chains_list = list(chains or [])
    footprints_list = list(footprints or [])

    wb = Workbook()
    # openpyxl always creates one default sheet; rename it to 'overview'
    # so we don't have an empty 'Sheet' kicking around.
    wb.active.title = "overview"

    # Each non-overview sheet writes its CSV-shaped data.
    csv_builders = [
        ("summary", chains_to_summary_csv(
            chains_list, container_type=container_type,
        )),
        ("exposure", chains_to_exposure_csv(
            chains_list, container_type=container_type,
        )),
        ("footprint", footprints_to_csv(
            footprints_list, container_type=container_type,
        )),
        ("regional", chains_to_regional_rollup_csv(
            chains_list, container_type=container_type,
        )),
        ("watchlist", chains_to_deficit_watchlist_csv(
            chains_list,
            container_type=container_type,
            threshold_days=threshold_days,
        )),
    ]

    sheet_rows: dict[str, int] = {}
    for sheet_name, csv_text in csv_builders:
        ws = wb.create_sheet(title=sheet_name)
        sheet_rows[sheet_name] = _write_csv_to_sheet(ws, csv_text)

    # Now write the overview sheet last (so the row counts are known).
    _write_overview_sheet(
        wb["overview"],
        container_type=container_type,
        threshold_days=threshold_days,
        sheet_rows=sheet_rows,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
