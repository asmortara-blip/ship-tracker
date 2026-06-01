"""Tests for utils.excel_export — multi-sheet xlsx workbook builder.

We never compare pixel-level styling. We pin structural facts:
  - openpyxl-only: the entire module is skipped if openpyxl is missing.
  - sheet names appear in workbook in expected order for each subset
  - empty data yields a minimal-but-valid workbook (no crash)
  - populated data writes headers + rows; data row count matches inputs
  - placeholder row fires on empty payloads
  - title + footer rows are present (text-level check)
  - cell colorization: positive numbers → green font, negative → red
  - macro indicator direction arrow ▲ / ▼ / — driven by delta sign
  - sentiment label thresholds: >0.15 Bullish, <-0.15 Bearish, else Neutral
  - summary BDI fallback key precedence
  - summary sentiment computed from list-of-dicts (avg score → label)
  - file_name suffix in get_excel_download_button_args includes report_type
  - legacy single-sheet helpers wrap export_to_excel correctly
  - export_to_excel("garbage_subset") falls back to "full"
  - export_to_excel(None) yields a workbook (None coerced to {})
  - workbook bytes are non-empty and reload via openpyxl.load_workbook

Covered surface
===============
- _style_header_row, _style_data_row (positive/negative/zero/string branches)
- _auto_fit_columns (clamps to [min_width, max_width])
- _add_sheet_title, _add_footer (text + merged cells)
- _write_table (header + alternating data rows; row index correct)
- _placeholder_row (empty-data branch)
- _workbook_to_bytes (round-trip)
- _build_summary, _build_freight_rates, _build_market_signals,
  _build_port_intelligence, _build_macro_indicators, _build_stock_performance,
  _build_news_sentiment
- export_to_excel (per subset; bad subset → full; None data; mixed builder errors)
- get_excel_download_button_args (label/mime/file_name keys + timestamp shape)
- export_port_data / export_route_data / export_signal_data / export_full_report
"""
from __future__ import annotations

import io

import pytest

# Skip the whole module if openpyxl isn't available in this environment.
pytest.importorskip("openpyxl")

import openpyxl  # noqa: E402
from openpyxl import Workbook  # noqa: E402

from utils.excel_export import (  # noqa: E402
    PLATFORM_FOOTER,
    _SHEET_BUILDERS,
    _SUBSET_MAP,
    _add_footer,
    _add_sheet_title,
    _auto_fit_columns,
    _build_freight_rates,
    _build_macro_indicators,
    _build_market_signals,
    _build_news_sentiment,
    _build_port_intelligence,
    _build_stock_performance,
    _build_summary,
    _placeholder_row,
    _style_data_row,
    _style_header_row,
    _workbook_to_bytes,
    _write_table,
    export_full_report,
    export_port_data,
    export_route_data,
    export_signal_data,
    export_to_excel,
    get_excel_download_button_args,
)


# ─── Helpers ────────────────────────────────────────────────────────────────


def _fresh_ws(title: str = "Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws


def _column_values(ws, col: int) -> list:
    return [ws.cell(row=r, column=col).value for r in range(1, ws.max_row + 1)]


def _row_values(ws, row: int) -> list:
    return [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]


def _all_cell_text(ws) -> str:
    parts = []
    for row in ws.iter_rows(values_only=True):
        parts.extend(str(v) for v in row if v is not None)
    return " | ".join(parts)


def _reload(wb_bytes: bytes):
    """Round-trip raw bytes back through openpyxl."""
    return openpyxl.load_workbook(io.BytesIO(wb_bytes))


# ─── Catalog / registry sanity ──────────────────────────────────────────────


def test_sheet_builders_keys_match_expected_set() -> None:
    expected = {
        "Summary",
        "Freight Rates",
        "Market Signals",
        "Port Intelligence",
        "Macro Indicators",
        "Stock Performance",
        "News Sentiment",
    }
    assert set(_SHEET_BUILDERS) == expected


def test_subset_map_full_lists_every_builder_once() -> None:
    full = _SUBSET_MAP["full"]
    assert len(full) == len(set(full)) == len(_SHEET_BUILDERS)
    assert set(full) == set(_SHEET_BUILDERS.keys())


def test_subset_map_every_subset_includes_summary_first() -> None:
    for name, sheets in _SUBSET_MAP.items():
        assert sheets[0] == "Summary", f"subset {name!r} doesn't start with Summary"


def test_subset_map_keys_are_documented_set() -> None:
    expected = {"full", "freight", "port", "stock", "macro", "sentiment", "signals"}
    assert set(_SUBSET_MAP) == expected


# ─── Style helpers ──────────────────────────────────────────────────────────


def test_style_header_row_paints_navy_fill_and_bold_white_font() -> None:
    _wb, ws = _fresh_ws()
    ws.append(["A", "B", "C"])
    _style_header_row(ws, 1)
    for c in ws[1]:
        assert c.font.bold is True
        # White hex in openpyxl Color object can be padded to 00FFFFFF
        assert (c.font.color.rgb or "").endswith("FFFFFF")
        assert (c.fill.fgColor.rgb or "").endswith("0F285A")


@pytest.mark.parametrize(
    "value,expected_color_suffix",
    [
        (42,    "0A6B3C"),    # positive → green
        (-3.14, "B91C1C"),    # negative → red
        (0,     "0F285A"),    # zero → navy
    ],
)
def test_style_data_row_numeric_colorization_by_sign(value, expected_color_suffix) -> None:
    _wb, ws = _fresh_ws()
    ws.append([value])
    _style_data_row(ws, 1, alt=False)
    cell = ws.cell(row=1, column=1)
    assert (cell.font.color.rgb or "").endswith(expected_color_suffix)


def test_style_data_row_string_value_uses_left_alignment() -> None:
    _wb, ws = _fresh_ws()
    ws.append(["hello"])
    _style_data_row(ws, 1, alt=False)
    cell = ws.cell(row=1, column=1)
    assert cell.alignment.horizontal == "left"


@pytest.mark.parametrize("alt,expected_fill", [(True, "F0F2F6"), (False, "FFFFFF")])
def test_style_data_row_alt_flag_swaps_fill(alt, expected_fill) -> None:
    _wb, ws = _fresh_ws()
    ws.append(["x"])
    _style_data_row(ws, 1, alt=alt)
    cell = ws.cell(row=1, column=1)
    assert (cell.fill.fgColor.rgb or "").endswith(expected_fill)


# ─── Column sizing + freeze ─────────────────────────────────────────────────


@pytest.mark.parametrize(
    "content,min_w,max_w,assert_fn",
    [
        ("x",       9, 50, lambda w: w >= 9),       # short → clamped to min
        ("x" * 500, 9, 50, lambda w: w <= 50),      # long  → clamped to max
    ],
)
def test_auto_fit_columns_respects_min_and_max_bounds(content, min_w, max_w, assert_fn) -> None:
    _wb, ws = _fresh_ws()
    ws.append([content])
    _auto_fit_columns(ws, min_width=min_w, max_width=max_w)
    assert assert_fn(ws.column_dimensions["A"].width)


# ─── Title / footer / placeholder / table ───────────────────────────────────


def test_add_sheet_title_writes_title_in_row_1() -> None:
    _wb, ws = _fresh_ws()
    _add_sheet_title(ws, "Hello World", ncols=3)
    assert ws.cell(row=1, column=1).value == "Hello World"


def test_add_footer_writes_platform_footer_string() -> None:
    _wb, ws = _fresh_ws()
    ws.append(["data"])  # row 1
    _add_footer(ws, row_num=3, ncols=4)
    text = _all_cell_text(ws)
    assert PLATFORM_FOOTER in text


def test_placeholder_row_writes_default_message() -> None:
    _wb, ws = _fresh_ws()
    _placeholder_row(ws, ncols=5)
    assert "No data available" in str(ws.cell(row=1, column=1).value)


def test_placeholder_row_accepts_custom_message() -> None:
    _wb, ws = _fresh_ws()
    _placeholder_row(ws, ncols=5, msg="Custom empty.")
    assert ws.cell(row=1, column=1).value == "Custom empty."


def test_write_table_appends_header_then_data_rows_and_returns_last_row_index() -> None:
    _wb, ws = _fresh_ws()
    headers = ["A", "B"]
    rows = [[1, 2], [3, 4], [5, 6]]
    last = _write_table(ws, headers, rows)
    # header at row 1, data at rows 2-4
    assert last == 4
    assert _row_values(ws, 1) == ["A", "B"]
    assert _row_values(ws, 4) == [5, 6]


def test_write_table_empty_rows_returns_header_row_index() -> None:
    _wb, ws = _fresh_ws()
    last = _write_table(ws, ["A", "B"], [])
    assert last == 1


# ─── Workbook bytes round-trip ──────────────────────────────────────────────


def test_workbook_to_bytes_produces_loadable_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hello"
    ws["A1"] = "world"
    raw = _workbook_to_bytes(wb)
    assert isinstance(raw, bytes) and len(raw) > 0
    rb = _reload(raw)
    assert rb["Hello"]["A1"].value == "world"


# ─── Builder: Summary ───────────────────────────────────────────────────────


def test_build_summary_writes_title_and_metric_rows_with_no_data() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Summary")
    _build_summary(ws, {})
    text = _all_cell_text(ws)
    assert "Market Summary Dashboard" in text
    assert "Baltic Dry Index (BDI)" in text
    assert PLATFORM_FOOTER in text
    # Metrics column has 10 rows of named labels
    metric_col = [v for v in _column_values(ws, 1) if v]
    assert "Report Generated" in metric_col
    assert "Platform" in metric_col


def test_build_summary_prefers_bdi_key_over_alternates() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Summary")
    _build_summary(ws, {"freight_data": {"bdi": 1234, "BDI": 9999}})
    text = _all_cell_text(ws)
    assert "1234" in text
    assert "9999" not in text


@pytest.mark.parametrize(
    "freight_payload,expected_token",
    [
        ({"BDI": 555}, "555"),
        ({"baltic_dry_index": 777}, "777"),
    ],
)
def test_build_summary_bdi_fallback_keys(freight_payload, expected_token) -> None:
    wb = Workbook()
    ws = wb.create_sheet("Summary")
    _build_summary(ws, {"freight_data": freight_payload})
    assert expected_token in _all_cell_text(ws)


@pytest.mark.parametrize(
    "scores,expected_label",
    [
        ([0.5, 0.3],   "Bullish"),
        ([-0.5, -0.4], "Bearish"),
        ([0.05, -0.05],"Neutral"),
    ],
)
def test_build_summary_sentiment_list_averages_to_label(scores, expected_label) -> None:
    wb = Workbook()
    ws = wb.create_sheet("Summary")
    _build_summary(ws, {"sentiment_data": [{"score": s} for s in scores]})
    assert expected_label in _all_cell_text(ws)


def test_build_summary_counts_ports_routes_signals() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Summary")
    _build_summary(
        ws,
        {
            "port_results":  [{"x": 1}, {"x": 2}, {"x": 3}],
            "route_results": [{"y": 1}, {"y": 2}],
            "stock_data":    [{"z": 1}, {"z": 2}, {"z": 3}, {"z": 4}],
            "insights":      [{}],
            "macro_data":    {"gdp": 1, "cpi": 2},
        },
    )
    vals = [v for v in _column_values(ws, 2) if v is not None]
    assert 3 in vals  # ports
    assert 2 in vals  # routes
    assert 4 in vals  # stocks


# ─── Builder: Freight rates ─────────────────────────────────────────────────


def test_build_freight_rates_empty_shows_placeholder() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Freight Rates")
    _build_freight_rates(ws, {})
    assert "No data available" in _all_cell_text(ws)


def test_build_freight_rates_populated_writes_headers_and_each_route() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Freight Rates")
    routes = [
        {"name": "Asia-Europe", "rate": 2500, "wow_change": 1.2, "yoy_change": -5.0},
        {"route": "Transpac", "current_rate": 3100},
    ]
    _build_freight_rates(ws, {"freight_data": {"routes": routes}})
    text = _all_cell_text(ws)
    assert "Asia-Europe" in text
    assert "Transpac" in text
    assert "Route Name" in text  # header


def test_build_freight_rates_falls_back_to_route_results_key() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Freight Rates")
    _build_freight_rates(ws, {"route_results": [{"name": "FBR"}]})
    assert "FBR" in _all_cell_text(ws)


# ─── Builder: Market signals ────────────────────────────────────────────────


def test_build_market_signals_empty_uses_placeholder() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Market Signals")
    _build_market_signals(ws, {})
    assert "No data available" in _all_cell_text(ws)


def test_build_market_signals_populated_lists_tickers() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Market Signals")
    _build_market_signals(ws, {
        "stock_data": [
            {"ticker": "ZIM", "signal_type": "BUY", "confidence": 0.85},
            {"symbol": "MAERSKB", "signal": "HOLD"},
        ]
    })
    text = _all_cell_text(ws)
    assert "ZIM" in text
    assert "MAERSKB" in text


# ─── Builder: Port intelligence ─────────────────────────────────────────────


def test_build_port_intelligence_populated_writes_port_rows() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Port Intelligence")
    _build_port_intelligence(ws, {
        "port_results": [
            {"name": "Singapore", "congestion": "Medium", "wait_time": 42, "throughput": 100000},
            {"port_name": "Rotterdam", "region": "Europe"},
        ]
    })
    text = _all_cell_text(ws)
    assert "Singapore" in text
    assert "Rotterdam" in text
    assert "Congestion Level" in text


def test_build_port_intelligence_empty_emits_placeholder_not_crash() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Port Intelligence")
    _build_port_intelligence(ws, {})
    assert "No data available" in _all_cell_text(ws)


# ─── Builder: Macro indicators ──────────────────────────────────────────────


def test_build_macro_indicators_dict_input_produces_one_row_per_key() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Macro")
    _build_macro_indicators(ws, {
        "macro_data": {
            "GDP":  {"current": 2.5, "prior": 2.1, "delta": 0.4},
            "CPI":  {"current": 3.1, "prior": 3.5, "delta": -0.4},
        }
    })
    text = _all_cell_text(ws)
    assert "GDP" in text
    assert "CPI" in text


@pytest.mark.parametrize(
    "delta,expected_glyph",
    [
        (1,     "▲"),
        (-1,    "▼"),
        ("n/a", "—"),
    ],
)
def test_build_macro_indicators_direction_glyph_by_delta(delta, expected_glyph) -> None:
    wb = Workbook()
    ws = wb.create_sheet("Macro")
    _build_macro_indicators(ws, {
        "macro_data": {"X": {"current": 8, "prior": 9, "delta": delta}}
    })
    assert expected_glyph in _all_cell_text(ws)


def test_build_macro_indicators_list_input_renders_rows() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Macro")
    _build_macro_indicators(ws, {
        "macro_data": [
            {"indicator": "PMI", "current": 50.1, "prior": 49.0, "delta": 1.1}
        ]
    })
    assert "PMI" in _all_cell_text(ws)


# ─── Builder: Stock performance ─────────────────────────────────────────────


def test_build_stock_performance_dict_input_converts_to_rows() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Stocks")
    _build_stock_performance(ws, {
        "stock_data": {
            "ZIM":  {"price": 12.5, "day_change": 0.5, "ytd": 25.0},
            "MAERSKB": 1250.0,
        }
    })
    text = _all_cell_text(ws)
    assert "ZIM" in text
    assert "MAERSKB" in text


def test_build_stock_performance_list_input_renders_companies() -> None:
    wb = Workbook()
    ws = wb.create_sheet("Stocks")
    _build_stock_performance(ws, {
        "stock_data": [{"ticker": "DAC", "company": "Danaos", "price": 80}]
    })
    text = _all_cell_text(ws)
    assert "DAC" in text
    assert "Danaos" in text


# ─── Builder: News sentiment ────────────────────────────────────────────────


@pytest.mark.parametrize(
    "score,expected_label",
    [
        ( 0.5,  "Bullish"),
        (-0.5,  "Bearish"),
        ( 0.05, "Neutral"),
        (-0.15, "Neutral"),   # boundary, not strictly < -0.15
    ],
)
def test_build_news_sentiment_label_thresholds(score, expected_label) -> None:
    wb = Workbook()
    ws = wb.create_sheet("News")
    _build_news_sentiment(ws, {"sentiment_data": [{"headline": "H1", "score": score}]})
    assert expected_label in _all_cell_text(ws)


def test_build_news_sentiment_truncates_long_headline_to_200_chars() -> None:
    wb = Workbook()
    ws = wb.create_sheet("News")
    long_headline = "Z" * 500
    _build_news_sentiment(ws, {"sentiment_data": [{"headline": long_headline, "score": 0}]})
    # Find the cell that contains the Z run
    found_lengths = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and v.startswith("ZZZZ"):
                found_lengths.append(len(v))
    assert found_lengths, "headline cell not found"
    assert max(found_lengths) == 200


def test_build_news_sentiment_dict_input_with_articles_key() -> None:
    wb = Workbook()
    ws = wb.create_sheet("News")
    _build_news_sentiment(ws, {
        "sentiment_data": {"articles": [{"headline": "Reefer demand rising", "score": 0.2}]}
    })
    assert "Reefer demand rising" in _all_cell_text(ws)


# ─── Public API: export_to_excel ────────────────────────────────────────────


def test_export_to_excel_full_produces_all_sheets_in_order() -> None:
    raw = export_to_excel({}, report_type="full")
    wb = _reload(raw)
    assert wb.sheetnames == list(_SUBSET_MAP["full"])


def test_export_to_excel_subset_freight_has_two_sheets() -> None:
    raw = export_to_excel({}, report_type="freight")
    wb = _reload(raw)
    assert wb.sheetnames == ["Summary", "Freight Rates"]


def test_export_to_excel_unknown_subset_falls_back_to_full() -> None:
    raw = export_to_excel({}, report_type="not_a_real_subset")
    wb = _reload(raw)
    assert wb.sheetnames == list(_SUBSET_MAP["full"])


def test_export_to_excel_none_data_coerces_to_empty_dict() -> None:
    raw = export_to_excel(None, report_type="port")
    wb = _reload(raw)
    assert "Summary" in wb.sheetnames
    assert "Port Intelligence" in wb.sheetnames


def test_export_to_excel_returns_non_empty_bytes() -> None:
    raw = export_to_excel({}, report_type="sentiment")
    assert isinstance(raw, bytes)
    assert len(raw) > 100  # xlsx has a minimum overhead


def test_export_to_excel_each_sheet_has_gridlines_disabled() -> None:
    raw = export_to_excel({}, report_type="full")
    wb = _reload(raw)
    for name in wb.sheetnames:
        assert wb[name].sheet_view.showGridLines is False


def test_export_to_excel_builder_exception_writes_error_cell_but_keeps_other_sheets(
    monkeypatch,
) -> None:
    from utils import excel_export as mod

    def boom(ws, data):
        raise ValueError("synthetic failure")

    monkeypatch.setitem(mod._SHEET_BUILDERS, "Summary", boom)
    raw = mod.export_to_excel({}, report_type="freight")
    wb = _reload(raw)
    # Summary still in workbook, but with the error placeholder
    assert "Summary" in wb.sheetnames
    assert "synthetic failure" in str(wb["Summary"].cell(row=3, column=1).value)
    # Other sheet still built normally
    assert "Freight Rates" in wb.sheetnames


# ─── Public API: get_excel_download_button_args ─────────────────────────────


def test_get_excel_download_button_args_returns_required_keys() -> None:
    args = get_excel_download_button_args({}, report_type="full")
    assert set(args.keys()) == {"label", "data", "file_name", "mime"}


def test_get_excel_download_button_args_file_name_includes_report_type() -> None:
    args = get_excel_download_button_args({}, report_type="signals")
    assert args["file_name"].startswith("shiptracker_signals_")
    assert args["file_name"].endswith(".xlsx")


def test_get_excel_download_button_args_uses_xlsx_mime() -> None:
    args = get_excel_download_button_args({}, report_type="freight")
    assert args["mime"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_get_excel_download_button_args_data_field_is_bytes() -> None:
    args = get_excel_download_button_args({}, report_type="macro")
    assert isinstance(args["data"], bytes) and len(args["data"]) > 100


# ─── Legacy single-sheet helpers ────────────────────────────────────────────


def test_export_port_data_builds_port_subset() -> None:
    raw = export_port_data([{"name": "Singapore"}])
    wb = _reload(raw)
    assert wb.sheetnames == ["Summary", "Port Intelligence"]
    assert "Singapore" in _all_cell_text(wb["Port Intelligence"])


def test_export_route_data_builds_freight_subset() -> None:
    raw = export_route_data([{"name": "TPEB"}], {"bdi": 1500})
    wb = _reload(raw)
    assert wb.sheetnames == ["Summary", "Freight Rates"]
    assert "1500" in _all_cell_text(wb["Summary"])


def test_export_signal_data_builds_signals_subset() -> None:
    raw = export_signal_data([{"ticker": "ZIM"}])
    wb = _reload(raw)
    assert wb.sheetnames == ["Summary", "Market Signals"]
    assert "ZIM" in _all_cell_text(wb["Market Signals"])


def test_export_full_report_uses_signals_when_provided() -> None:
    raw = export_full_report(
        port_results=[],
        route_results=[],
        insights=[],
        freight_data={},
        macro_data={},
        stock_data=[{"ticker": "OTHER"}],
        signals=[{"ticker": "WINNER"}],
    )
    wb = _reload(raw)
    # signals override stock_data per source contract
    txt = _all_cell_text(wb["Stock Performance"])
    assert "WINNER" in txt
    assert "OTHER" not in txt


def test_export_full_report_falls_back_to_stock_data_when_signals_none() -> None:
    raw = export_full_report(
        port_results=[],
        route_results=[],
        insights=[],
        freight_data={},
        macro_data={},
        stock_data=[{"ticker": "FALLBACK"}],
        signals=None,
    )
    wb = _reload(raw)
    assert "FALLBACK" in _all_cell_text(wb["Stock Performance"])
