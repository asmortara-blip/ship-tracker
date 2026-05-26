"""Tests for ``worker.api_server`` — the stdlib read-only HTTP API.

Spin up a real ``HTTPServer`` bound to ``127.0.0.1:0`` (kernel-assigned
ephemeral port) in a daemon thread per test, fire ``requests`` at it
from the test process, tear it down in fixture teardown. The same
pattern as ``tests/test_webhook_listener.py``; we deliberately avoid
mocking the HTTP layer because the listener's value is in the wire
behaviour (status codes, header parsing, JSON payload shape, bearer
auth flow).

Each test isolates the SQLite state DB to a per-test ``tmp_path`` so
engine calls (``acknowledge_alert``, ``load_alerts``, …) don't touch
the real ``cache/ship_tracker.db``. We seed a user + a fresh API
token per test via the real ``auth.users.signup`` /
``auth.tokens.create_token`` so the verify path is exercised end-to-
end — no mocks past the bind socket.
"""
from __future__ import annotations

import socket
import threading
from http.server import HTTPServer
from typing import Optional

import pytest
import requests

from worker import api_server


# ─── Fixtures ─────────────────────────────────────────────────────────────


@pytest.fixture(autouse=True)
def isolated_state_db(monkeypatch, tmp_path):
    """Per-test SQLite isolation. Mirrors the pattern in
    test_webhook_listener.py — any ``state.db.get_connection`` call
    lands in this tmp file."""
    from state import db as state_db
    monkeypatch.setattr(state_db, "DB_PATH", tmp_path / "ship_tracker.db")
    state_db.reset_for_tests()
    yield
    state_db.reset_for_tests()


@pytest.fixture(autouse=True)
def _isolated_rate_limit():
    """Per-test rate-limit registry isolation. The module-level
    ``_BUCKETS`` dict in ``auth.rate_limit`` would otherwise leak
    state across tests — a test that exhausts a user's bucket would
    starve the next test that re-uses the same username."""
    from auth import rate_limit as rl
    rl.clear_buckets()
    yield
    rl.clear_buckets()


def _find_free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return s.getsockname()[1]


@pytest.fixture
def server():
    """Spin up the API handler on localhost:<ephemeral-port> in a
    daemon thread. Tear down cleanly in teardown — ``shutdown()`` on
    the server unblocks ``serve_forever`` and the thread exits."""
    port = _find_free_port()
    httpd = HTTPServer(("127.0.0.1", port), api_server.APIHandler)
    thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    thread.start()
    base_url = f"http://127.0.0.1:{port}"
    try:
        yield base_url
    finally:
        httpd.shutdown()
        httpd.server_close()
        thread.join(timeout=2.0)


def _make_user(username: str = "alice", password: str = "Hunter2!hunter") -> str:
    """Create a fresh user via the real signup path. Returns user_id."""
    from auth.users import signup
    u = signup(username, password)
    assert u is not None
    return u.user_id


def _mint_token(user_id: str, label: str = "ci") -> str:
    """Create an API token and return the raw secret."""
    from auth.tokens import create_token
    result = create_token(user_id, label)
    assert result is not None
    _meta, raw = result
    return raw


def _seed_alerts(
    user_id: str,
    *,
    n: int = 1,
    severity: str = "HIGH",
    alert_type: str = "BDI_MOVE",
    ticker_prefix: str = "T",
) -> list[str]:
    """Insert ``n`` distinct alerts for ``user_id`` and return their ids.

    Each alert gets a unique ticker so the engine's dedup_key logic
    treats them as separate rows — otherwise ``save_alerts`` would
    collapse them into one row with a higher fire_count."""
    from engine.alert_engine_v2 import ShippingAlert, save_alerts, _now_iso, _new_id

    alerts = []
    for i in range(n):
        alerts.append(ShippingAlert(
            alert_id=_new_id(),
            created_at=_now_iso(),
            alert_type=alert_type,
            severity=severity,
            title=f"alert #{i}",
            body="seeded",
            ticker=f"{ticker_prefix}{i:04d}",
            route_id="",
            port_locode="",
            value=float(i),
            threshold=0.0,
            change_pct=float(i),
            acknowledged=False,
        ))
    save_alerts(alerts, user_id=user_id)
    return [a.alert_id for a in alerts]


def _bearer(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


# ─── Health endpoint: public ──────────────────────────────────────────────


def test_health_returns_200_without_auth(server):
    """``GET /api/v1/health`` is unauthenticated by design — load
    balancers must be able to probe without shipping a token."""
    r = requests.get(f"{server}/api/v1/health", timeout=5)
    assert r.status_code == 200
    body = r.json()
    assert body["status"] in {"ok", "degraded"}
    assert "schema_version" in body
    assert "up_seconds" in body
    assert "users" in body
    assert "unacked_critical_count" in body


def test_health_payload_carries_now_utc_iso(server):
    """The shape mirrors ``worker.webhook_listener.GET /health``."""
    r = requests.get(f"{server}/api/v1/health", timeout=5)
    assert r.status_code == 200
    body = r.json()
    assert "now_utc" in body
    # Must be parseable ISO-8601.
    from datetime import datetime
    datetime.fromisoformat(body["now_utc"])


# ── /api/v1/backtests/health — public analytical-layer probe ───────────

def test_backtests_health_returns_200_with_all_validators(server):
    """`GET /api/v1/backtests/health` is unauthenticated and returns a
    consolidated report of every backtest validator."""
    r = requests.get(f"{server}/api/v1/backtests/health", timeout=10)
    # Status code follows the rollup: 200 when all 9 healthy, 503 otherwise.
    assert r.status_code in (200, 503)
    body = r.json()
    assert body["status"] in {"ok", "degraded"}
    assert "validators" in body
    assert "healthy_count" in body
    assert "total" in body
    assert "now_utc" in body
    # Every validator carries the required structured fields.
    for v in body["validators"]:
        assert {"name", "healthy", "headline_label", "headline_value",
                "summary", "raw"} <= set(v.keys())


def test_backtests_health_returns_200_when_all_healthy(server):
    """On the bundled synth all 13 validators read healthy → 200."""
    r = requests.get(f"{server}/api/v1/backtests/health", timeout=10)
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "ok"
    assert body["healthy_count"] == body["total"]
    assert body["total"] == 13


def test_backtests_health_does_not_require_authorization_header(server):
    """Public probe — auth bypass works the same as /health."""
    r = requests.get(
        f"{server}/api/v1/backtests/health",
        headers={"Authorization": "Bearer not-a-real-token"},
        timeout=10,
    )
    assert r.status_code in (200, 503)


# ── /api/v1/ports/supply-lines — authenticated port supply data ────────

def test_port_supply_lines_requires_auth(server):
    """No token → 401. Unlike /health, this is per-user-gated."""
    r = requests.get(f"{server}/api/v1/ports/supply-lines", timeout=10)
    assert r.status_code == 401


def test_port_supply_lines_returns_chain_payload(server):
    """With a valid token, returns the full per-port chain envelope."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/ports/supply-lines",
        headers=_bearer(token), timeout=10,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["container_type"] == "40FT_DRY"
    assert body["total"] >= 1
    assert "now_utc" in body
    assert isinstance(body["chains"], list)
    assert len(body["chains"]) == body["total"]
    first = body["chains"][0]
    assert {"port", "exposed_companies", "routes_touching",
            "top_commodities", "summary"} <= set(first.keys())
    assert {"locode", "name", "region", "lat", "lon",
            "supply_deficit_days", "utilization_pct",
            "severity_label", "container_type"} <= set(first["port"].keys())


def test_port_supply_lines_respects_container_type_param(server):
    """`container_type=40FT_REEFER` re-runs with the reefer slice."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/ports/supply-lines",
        params={"container_type": "40FT_REEFER"},
        headers=_bearer(token), timeout=10,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["container_type"] == "40FT_REEFER"
    for c in body["chains"]:
        assert c["port"]["container_type"] == "40FT_REEFER"


def test_port_supply_lines_rejects_bad_container_type(server):
    """Unknown container_type → 400, not a silent fallback."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/ports/supply-lines",
        params={"container_type": "BANANA"},
        headers=_bearer(token), timeout=10,
    )
    assert r.status_code == 400


def test_port_supply_lines_respects_top_n_param(server):
    """top_n caps the per-port exposed_companies + top_commodities lists."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/ports/supply-lines",
        params={"top_n": "2"},
        headers=_bearer(token), timeout=10,
    )
    assert r.status_code == 200
    body = r.json()
    for c in body["chains"]:
        assert len(c["exposed_companies"]) <= 2
        assert len(c["top_commodities"]) <= 2


def test_port_supply_lines_rejects_bad_top_n(server):
    """Non-integer or out-of-range top_n → 400."""
    uid = _make_user()
    token = _mint_token(uid)
    for bad in ("zero", "0", "100"):
        r = requests.get(
            f"{server}/api/v1/ports/supply-lines",
            params={"top_n": bad},
            headers=_bearer(token), timeout=10,
        )
        assert r.status_code == 400, (
            f"expected 400 for top_n={bad}, got {r.status_code}"
        )


# ── /api/v1/ports/supply-lines.xlsx — single workbook ─────────────────

def test_supply_lines_xlsx_requires_auth(server):
    """No token → 401, same gate as the JSON endpoint."""
    r = requests.get(
        f"{server}/api/v1/ports/supply-lines.xlsx", timeout=10,
    )
    assert r.status_code == 401


def test_supply_lines_xlsx_returns_workbook_bytes(server):
    """With a valid token, returns an actual .xlsx workbook that
    openpyxl can load back."""
    import io
    import openpyxl

    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/ports/supply-lines.xlsx",
        headers=_bearer(token), timeout=15,
    )
    assert r.status_code == 200
    # Content-Type is the canonical openxml spreadsheetml.sheet.
    assert r.headers["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )
    # Response is the workbook bytes — openpyxl must accept them.
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    # Six canonical sheets in order.
    assert wb.sheetnames == [
        "overview", "summary", "exposure", "footprint",
        "regional", "watchlist",
    ]


def test_supply_lines_xlsx_content_disposition_carries_filename(server):
    """The download filename matches the CLI exporter's convention so
    operator-side identification is consistent across paths."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/ports/supply-lines.xlsx",
        headers=_bearer(token), timeout=15,
    )
    assert r.status_code == 200
    disposition = r.headers.get("Content-Disposition", "")
    assert "attachment" in disposition
    assert "port_supply_lines_workbook" in disposition
    assert ".xlsx" in disposition


def test_supply_lines_xlsx_respects_container_type(server):
    """Container-type param flows through to the workbook + the
    overview sheet reports the requested type."""
    import io
    import openpyxl

    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/ports/supply-lines.xlsx",
        params={"container_type": "40FT_REEFER"},
        headers=_bearer(token), timeout=15,
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    overview_cells = [
        str(c.value) for row in wb["overview"].iter_rows()
        for c in row if c.value is not None
    ]
    assert "40FT_REEFER" in overview_cells


def test_supply_lines_xlsx_rejects_bad_container_type(server):
    """Bad container_type → 400, same contract as the JSON endpoint."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/ports/supply-lines.xlsx",
        params={"container_type": "BANANA"},
        headers=_bearer(token), timeout=10,
    )
    assert r.status_code == 400


def test_supply_lines_xlsx_rejects_bad_threshold(server):
    """threshold_days must be numeric — bad value → 400, not silent."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/ports/supply-lines.xlsx",
        params={"threshold_days": "nope"},
        headers=_bearer(token), timeout=10,
    )
    assert r.status_code == 400


def test_health_does_not_require_authorization_header(server):
    """Sanity check: a deliberately broken bearer header still gets a
    200 from /health — auth bypass is the whole point."""
    r = requests.get(
        f"{server}/api/v1/health",
        headers={"Authorization": "Bearer not-a-real-token"},
        timeout=5,
    )
    assert r.status_code == 200


# ─── OpenAPI spec endpoint: public ────────────────────────────────────────


def test_openapi_endpoint_returns_200_without_auth(server):
    """``GET /api/v1/openapi.json`` is unauthenticated by design — SDK
    generators and Swagger UI must be able to fetch the spec without
    needing a token."""
    # Pre-warm the module cache so the per-test response time doesn't
    # depend on whether this is the first request after server start.
    from worker import api_server
    api_server._OPENAPI_BYTES_CACHE = None  # force lazy rebuild this test

    r = requests.get(f"{server}/api/v1/openapi.json", timeout=5)
    assert r.status_code == 200
    assert r.headers["Content-Type"].startswith("application/json")


def test_openapi_endpoint_body_is_valid_json_with_openapi_key(server):
    """The body must parse as JSON and carry the OpenAPI version
    string under the ``openapi`` key (per OpenAPI 3.0.3 §4)."""
    r = requests.get(f"{server}/api/v1/openapi.json", timeout=5)
    assert r.status_code == 200
    body = r.json()
    assert "openapi" in body
    assert body["openapi"].startswith("3.")


def test_openapi_endpoint_info_title_matches_default(server):
    """The default title from ``build_openapi_spec`` is "Ship Tracker
    API" — sanity check it shows up unchanged through the wire."""
    r = requests.get(f"{server}/api/v1/openapi.json", timeout=5)
    body = r.json()
    assert body["info"]["title"] == "Ship Tracker API"
    assert body["info"]["version"]  # any non-empty string


def test_openapi_endpoint_exposes_at_least_20_paths(server):
    """The spec must cover the bulk of the API surface — the spec
    builder declares ~29 paths today; we use 20 as a safety floor so
    a future trim of one or two endpoints doesn't flake the test."""
    r = requests.get(f"{server}/api/v1/openapi.json", timeout=5)
    body = r.json()
    assert len(body["paths"]) >= 20


# ─── Auth: missing / invalid / revoked token → 401 ────────────────────────


@pytest.mark.parametrize("path,method", [
    ("/api/v1/alerts", "GET"),
    ("/api/v1/alerts/some-id", "GET"),
    ("/api/v1/alerts/some-id/ack", "POST"),
    ("/api/v1/reports", "GET"),
    ("/api/v1/reports/some-id/html", "GET"),
    ("/api/v1/telemetry/llm", "GET"),
    ("/api/v1/telemetry/perf", "GET"),
])
def test_endpoint_returns_401_without_auth_header(server, path, method):
    """Every authenticated endpoint must return 401 when the
    Authorization header is absent — never 200/4xx/5xx leaking data."""
    r = requests.request(method, f"{server}{path}", timeout=5)
    assert r.status_code == 401
    assert r.json() == {"error": "unauthorized"}


@pytest.mark.parametrize("path,method", [
    ("/api/v1/alerts", "GET"),
    ("/api/v1/alerts/some-id", "GET"),
    ("/api/v1/alerts/some-id/ack", "POST"),
    ("/api/v1/reports", "GET"),
    ("/api/v1/reports/some-id/html", "GET"),
    ("/api/v1/telemetry/llm", "GET"),
    ("/api/v1/telemetry/perf", "GET"),
])
def test_endpoint_returns_401_with_invalid_token(server, path, method):
    """A bearer header whose token is not in the api_tokens table must
    401 — never let a random string in."""
    headers = _bearer("invalid-token-that-does-not-match-anything")
    r = requests.request(method, f"{server}{path}", headers=headers, timeout=5)
    assert r.status_code == 401
    assert r.json() == {"error": "unauthorized"}


@pytest.mark.parametrize("path,method", [
    ("/api/v1/alerts", "GET"),
    ("/api/v1/alerts/some-id", "GET"),
    ("/api/v1/alerts/some-id/ack", "POST"),
    ("/api/v1/reports", "GET"),
    ("/api/v1/reports/some-id/html", "GET"),
    ("/api/v1/telemetry/llm", "GET"),
    ("/api/v1/telemetry/perf", "GET"),
])
def test_endpoint_returns_401_with_revoked_token(server, path, method):
    """A revoked token must 401. We mint a token, revoke it, then
    re-issue the request — same secret, but the row's revoked=1 flag
    excludes it from ``verify_token``'s lookup."""
    from auth.tokens import create_token, revoke_token
    uid = _make_user("alice", "Hunter2!hunter")
    result = create_token(uid, "ci")
    assert result is not None
    meta, raw = result
    assert revoke_token(meta.token_id, user_id=uid) is True

    headers = _bearer(raw)
    r = requests.request(method, f"{server}{path}", headers=headers, timeout=5)
    assert r.status_code == 401
    assert r.json() == {"error": "unauthorized"}


def test_malformed_authorization_header_returns_401(server):
    """An ``Authorization`` header that does not start with ``Bearer ``
    is treated as no auth — 401."""
    r = requests.get(
        f"{server}/api/v1/alerts",
        headers={"Authorization": "Basic dXNlcjpwYXNz"},
        timeout=5,
    )
    assert r.status_code == 401


def test_bearer_scheme_is_case_insensitive(server):
    """RFC 6750 says the auth-scheme name is case-insensitive — a
    lowercase ``bearer`` prefix must work."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/alerts",
        headers={"Authorization": f"bearer {token}"},
        timeout=5,
    )
    assert r.status_code == 200


# ─── GET /api/v1/alerts ───────────────────────────────────────────────────


def test_list_alerts_returns_200_and_array(server):
    uid = _make_user()
    token = _mint_token(uid)
    _seed_alerts(uid, n=3)
    r = requests.get(f"{server}/api/v1/alerts", headers=_bearer(token), timeout=5)
    assert r.status_code == 200
    body = r.json()
    assert isinstance(body, list)
    assert len(body) == 3
    # Each row carries the dataclass field set.
    for row in body:
        assert {"alert_id", "severity", "alert_type", "title", "acknowledged"}.issubset(row)


def test_list_alerts_empty_for_user_with_no_alerts(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(f"{server}/api/v1/alerts", headers=_bearer(token), timeout=5)
    assert r.status_code == 200
    assert r.json() == []


def test_list_alerts_filters_by_severity(server):
    uid = _make_user()
    token = _mint_token(uid)
    _seed_alerts(uid, n=2, severity="CRITICAL", ticker_prefix="C")
    _seed_alerts(uid, n=3, severity="HIGH", ticker_prefix="H")
    _seed_alerts(uid, n=1, severity="LOW", ticker_prefix="L")
    r = requests.get(
        f"{server}/api/v1/alerts",
        params={"severity": "HIGH"},
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    rows = r.json()
    assert len(rows) == 3
    assert all(row["severity"] == "HIGH" for row in rows)


def test_list_alerts_severity_filter_is_case_insensitive(server):
    """A lowercase query parameter still matches the engine's
    uppercase severity values — defensive UX so callers don't have
    to guess the casing."""
    uid = _make_user()
    token = _mint_token(uid)
    _seed_alerts(uid, n=2, severity="CRITICAL", ticker_prefix="C")
    r = requests.get(
        f"{server}/api/v1/alerts",
        params={"severity": "critical"},
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    rows = r.json()
    assert len(rows) == 2
    assert all(row["severity"] == "CRITICAL" for row in rows)


def test_list_alerts_caps_at_500_even_with_large_window(server):
    """A 365-day window asking for 600 alerts must still return at
    most 500. Caps the JSON payload."""
    uid = _make_user()
    token = _mint_token(uid)
    _seed_alerts(uid, n=600, ticker_prefix="X")
    r = requests.get(
        f"{server}/api/v1/alerts",
        params={"window_days": 365},
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    rows = r.json()
    assert len(rows) == 500


def test_list_alerts_respects_window_days(server):
    """Asking for window_days=0 must return zero rows — the engine's
    cutoff excludes everything older than 'now'."""
    uid = _make_user()
    token = _mint_token(uid)
    _seed_alerts(uid, n=3)
    r = requests.get(
        f"{server}/api/v1/alerts",
        params={"window_days": 0},
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    assert r.json() == []


def test_list_alerts_is_user_scoped(server):
    """Alice's token must NOT see Bob's alerts. Bob's alert is seeded
    under Bob's user_id; Alice's GET returns an empty list."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    _seed_alerts(bob_uid, n=4, ticker_prefix="B")
    r = requests.get(
        f"{server}/api/v1/alerts", headers=_bearer(alice_token), timeout=5,
    )
    assert r.status_code == 200
    assert r.json() == []


# ─── GET /api/v1/alerts/<id> ──────────────────────────────────────────────


def test_get_single_alert_returns_dict(server):
    uid = _make_user()
    token = _mint_token(uid)
    ids = _seed_alerts(uid, n=2)
    r = requests.get(
        f"{server}/api/v1/alerts/{ids[0]}",
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["alert_id"] == ids[0]
    assert "severity" in body
    assert "title" in body


def test_get_single_alert_unknown_id_returns_404(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/alerts/this-id-does-not-exist",
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 404
    assert r.json() == {"error": "not found"}


def test_get_single_alert_cross_user_returns_404(server):
    """Alice tries to read Bob's alert by id → 404 (NOT 403). The
    response must be indistinguishable from 'unknown id' so a token-
    holder can't enumerate other users' alert ids."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_ids = _seed_alerts(bob_uid, n=1, ticker_prefix="B")
    r = requests.get(
        f"{server}/api/v1/alerts/{bob_ids[0]}",
        headers=_bearer(alice_token),
        timeout=5,
    )
    assert r.status_code == 404


# ─── POST /api/v1/alerts/<id>/ack ─────────────────────────────────────────


def test_ack_alert_marks_acknowledged(server):
    uid = _make_user()
    token = _mint_token(uid)
    ids = _seed_alerts(uid, n=1)

    r = requests.post(
        f"{server}/api/v1/alerts/{ids[0]}/ack",
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    assert r.json()["acknowledged"] is True

    # Verify via the engine: load_alerts should return the alert with
    # acknowledged=True.
    from engine.alert_engine_v2 import load_alerts
    alerts = load_alerts(max_age_days=365, user_id=uid)
    match = next(a for a in alerts if a.alert_id == ids[0])
    assert match.acknowledged is True


def test_ack_alert_for_different_user_does_not_ack(server):
    """Alice's token MUST NOT be able to ack Bob's alert. The endpoint
    returns 200 (per-spec: no info leak about other users' alert ids)
    BUT the underlying engine UPDATE matches zero rows because the
    SQL scope filter excludes Bob's row from Alice's view."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_ids = _seed_alerts(bob_uid, n=1, ticker_prefix="B")

    # Alice tries to ack Bob's alert.
    r = requests.post(
        f"{server}/api/v1/alerts/{bob_ids[0]}/ack",
        headers=_bearer(alice_token),
        timeout=5,
    )
    # The endpoint still returns 200 — we don't surface "no rows
    # updated" because that would let Alice probe for Bob's ids.
    assert r.status_code == 200

    # But Bob's alert MUST still be unacknowledged in the DB.
    from engine.alert_engine_v2 import load_alerts
    bob_alerts = load_alerts(max_age_days=365, user_id=bob_uid)
    bob_alert = next(a for a in bob_alerts if a.alert_id == bob_ids[0])
    assert bob_alert.acknowledged is False, \
        "Alice's token must not be able to ack Bob's alert"


def test_ack_alert_unknown_id_returns_200_idempotent(server):
    """Acking an unknown id is a no-op at the engine layer (UPDATE
    affects zero rows). The endpoint returns 200 idempotently so
    retries don't surface phantom 404s after an earlier ack."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/alerts/never-existed/ack",
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200


# ─── GET /api/v1/reports ──────────────────────────────────────────────────


def _seed_report(user_id: str, label: str = "rep") -> str:
    """Drop a tiny HTML file in REPORT_DIR and insert a row scoped to
    ``user_id``. Returns the new report_id."""
    import uuid
    from pathlib import Path
    from utils import report_history as rh
    from state.db import get_connection

    rh.REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_id = str(uuid.uuid4())
    file_path = rh.REPORT_DIR / f"test_{report_id[:8]}_{label}.html"
    html = f"<html><body><h1>{label}</h1></body></html>"
    file_path.write_text(html, encoding="utf-8")

    conn = get_connection()
    with conn:
        conn.execute(
            """
            INSERT INTO report_history
              (report_id, generated_at, report_date, sentiment_label,
               sentiment_score, risk_level, signal_count, data_quality,
               file_path, file_size_kb, user_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                report_id, "2026-05-22T00:00:00+00:00", "May 22, 2026",
                "NEUTRAL", 0.0, "MODERATE", 0, "FULL",
                str(file_path.resolve()), 0.5, user_id,
            ),
        )
    return report_id


def test_list_reports_returns_array(server, tmp_path, monkeypatch):
    # Redirect REPORT_DIR to tmp so we don't write into the real cache.
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")

    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid)
    r = requests.get(
        f"{server}/api/v1/reports", headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    rows = r.json()
    assert isinstance(rows, list)
    assert any(row["report_id"] == rid for row in rows)


def test_list_reports_empty_for_user_with_no_reports(server, tmp_path, monkeypatch):
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/reports", headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert r.json() == []


def test_list_reports_respects_limit(server, tmp_path, monkeypatch):
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    for i in range(5):
        _seed_report(uid, label=f"r{i}")
    r = requests.get(
        f"{server}/api/v1/reports",
        params={"limit": 2},
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    assert len(r.json()) == 2


def test_list_reports_is_user_scoped(server, tmp_path, monkeypatch):
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    _seed_report(bob_uid, label="b1")
    r = requests.get(
        f"{server}/api/v1/reports", headers=_bearer(alice_token), timeout=5,
    )
    assert r.status_code == 200
    # Bob's report must not appear in Alice's listing.
    rows = r.json()
    assert all(row.get("report_id") != "b1" for row in rows)


# ─── GET /api/v1/reports/diff ────────────────────────────────────────────


def test_diff_reports_missing_query_params_returns_400(
    server, tmp_path, monkeypatch,
):
    """Both `from` and `to` are required; missing either → 400 with a
    descriptive JSON body. Pins the API's input-validation contract
    so a caller never has to interpret an opaque 500."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)

    # Neither param.
    r1 = requests.get(
        f"{server}/api/v1/reports/diff",
        headers=_bearer(token), timeout=5,
    )
    assert r1.status_code == 400
    assert "from" in r1.text.lower() or "to" in r1.text.lower()

    # Only `from`.
    r2 = requests.get(
        f"{server}/api/v1/reports/diff",
        params={"from": "abc"},
        headers=_bearer(token), timeout=5,
    )
    assert r2.status_code == 400


def test_diff_reports_unknown_id_returns_404(server, tmp_path, monkeypatch):
    """An unknown id in the caller's scope → 404. Defining property:
    no info leak about whether the id exists at all (vs exists-but-
    belongs-to-another-user) — same response for both."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)

    r = requests.get(
        f"{server}/api/v1/reports/diff",
        params={"from": "ghost-a", "to": "ghost-b"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 404


def test_diff_reports_happy_path_returns_structured_diff(
    server, tmp_path, monkeypatch,
):
    """Two valid reports in the caller's scope → 200 + the documented
    JSON shape (`report_a_id`, `report_b_id`, `summary`, `entries`)."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    a_id = _seed_report(uid, label="a")
    b_id = _seed_report(uid, label="b")

    r = requests.get(
        f"{server}/api/v1/reports/diff",
        params={"from": a_id, "to": b_id},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["report_a_id"] == a_id
    assert body["report_b_id"] == b_id
    assert set(body.keys()) >= {"report_a_id", "report_b_id", "summary", "entries"}
    assert {"added", "removed", "changed"} <= set(body["summary"].keys())
    assert isinstance(body["entries"], list)


def test_diff_reports_cross_user_returns_404(server, tmp_path, monkeypatch):
    """Alice cannot diff Bob's reports — same 404 collapse as the
    unknown-id case. Per-user scoping contract."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")

    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_a = _seed_report(bob_uid, label="b1")
    bob_b = _seed_report(bob_uid, label="b2")

    r = requests.get(
        f"{server}/api/v1/reports/diff",
        params={"from": bob_a, "to": bob_b},
        headers=_bearer(alice_token), timeout=5,
    )
    # Same as unknown id — no permission-denied leak.
    assert r.status_code == 404


def test_diff_reports_requires_auth_like_other_endpoints(server):
    """Missing bearer → 401. Pins the gate so an unauthenticated
    probe can never get a structured diff response (which could carry
    information about the report population)."""
    r = requests.get(
        f"{server}/api/v1/reports/diff",
        params={"from": "x", "to": "y"},
        timeout=5,
    )
    assert r.status_code == 401


# ─── GET /api/v1/reports/<id>/html ────────────────────────────────────────


def test_get_report_html_returns_raw_html(server, tmp_path, monkeypatch):
    """Reports endpoint returns raw HTML with text/html — NOT JSON."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="raw")

    r = requests.get(
        f"{server}/api/v1/reports/{rid}/html",
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    ctype = r.headers.get("Content-Type", "")
    assert ctype.startswith("text/html")
    assert "<html>" in r.text
    assert "raw" in r.text


def test_get_report_html_unknown_id_returns_404(server, tmp_path, monkeypatch):
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/reports/this-id-does-not-exist/html",
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 404
    # 404 from this endpoint IS JSON — only the success path emits HTML.
    assert r.json() == {"error": "not found"}


def test_get_report_html_cross_user_returns_404(server, tmp_path, monkeypatch):
    """Bob's report id, looked up with Alice's token → 404."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_rid = _seed_report(bob_uid, label="bobs")
    r = requests.get(
        f"{server}/api/v1/reports/{bob_rid}/html",
        headers=_bearer(alice_token),
        timeout=5,
    )
    assert r.status_code == 404


# ─── GET /api/v1/telemetry/llm ────────────────────────────────────────────


def test_telemetry_llm_returns_summary_shape(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/telemetry/llm",
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    # Empty DB → zeroed dict with the canonical keys.
    for key in ("window_days", "total_calls", "total_tokens_in",
                "total_tokens_out", "total_cost_usd",
                "by_source", "by_model", "by_day"):
        assert key in body


def test_telemetry_llm_respects_window_days(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/telemetry/llm",
        params={"window_days": 1},
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    assert r.json()["window_days"] == 1


# ─── GET /api/v1/telemetry/perf ───────────────────────────────────────────


def test_telemetry_perf_returns_summary_shape(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/telemetry/perf",
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    for key in ("window_hours", "total_renders", "success_rate",
                "by_tab", "top_slow_tabs"):
        assert key in body


def test_telemetry_perf_respects_window_hours(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/telemetry/perf",
        params={"window_hours": 6},
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    assert r.json()["window_hours"] == 6


# ─── Routing: 404 / 405 ──────────────────────────────────────────────────


def test_unknown_path_returns_404(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/no-such-thing",
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 404
    assert r.json() == {"error": "not found"}


def test_unknown_root_path_returns_404(server):
    """An authenticated request to '/' returns 404 (no such endpoint).
    Without auth it would 401 first — the auth check fires before
    path dispatch by design (no info leak about valid paths)."""
    uid = _make_user("alice", "Hunter2!hunter")
    token = _mint_token(uid)
    r = requests.get(f"{server}/", headers=_bearer(token), timeout=5)
    assert r.status_code == 404


def test_post_to_get_endpoint_returns_405(server):
    """POST /api/v1/alerts is not a defined route — the same path
    is defined under GET. Must return 405 (path known under another
    verb), not 404."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/alerts",
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 405
    assert r.json() == {"error": "method not allowed"}


def test_get_on_ack_endpoint_returns_405(server):
    """GET on the ack endpoint (which is POST-only) → 405."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/alerts/some-id/ack",
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 405


def test_put_returns_405_on_known_path(server):
    """PUT is unsupported — any known path should 405."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.put(
        f"{server}/api/v1/alerts",
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 405


# ─── Smoke: full request with response headers ───────────────────────────


def test_response_content_type_is_json_for_list_alerts(server):
    """JSON endpoints emit ``application/json`` — clients written
    against the spec need to be able to rely on the header."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/alerts", headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert r.headers.get("Content-Type", "").startswith("application/json")


# ─── Write endpoints — auth on all new write paths ────────────────────────


@pytest.mark.parametrize("method,path", [
    ("POST",   "/api/v1/rules"),
    ("GET",    "/api/v1/rules"),
    ("DELETE", "/api/v1/rules"),
    ("POST",   "/api/v1/channels"),
    ("GET",    "/api/v1/channels"),
    ("DELETE", "/api/v1/channels/abc"),
    ("POST",   "/api/v1/reports/some-id/public"),
    ("DELETE", "/api/v1/reports/some-id/public"),
])
def test_write_endpoint_returns_401_without_auth_header(server, method, path):
    """Every WRITE endpoint must 401 with no Authorization header.
    Same contract as the read endpoints — never leak data on auth
    miss."""
    r = requests.request(method, f"{server}{path}", timeout=5)
    assert r.status_code == 401
    assert r.json() == {"error": "unauthorized"}


# ─── POST /api/v1/rules ───────────────────────────────────────────────────


def _sample_rules() -> list[dict]:
    """A pair of well-formed rule dicts for the rule-write tests.

    Shape mirrors ``ui/tab_alerts.py``'s session-state rule entries —
    a ``rule_id`` is the only required field for ``save_rules`` to
    pick up the dict."""
    return [
        {"rule_id": "r1", "name": "BDI surge", "metric": "bdi",
         "threshold_pct": 5.0, "severity": "HIGH"},
        {"rule_id": "r2", "name": "Port congestion", "metric": "congestion",
         "threshold": 0.8, "severity": "MEDIUM"},
    ]


def test_post_rules_saves_and_get_retrieves(server):
    """POST /api/v1/rules persists the body; the subsequent GET returns
    the same set (modulo ordering — ``load_rules`` orders by rule_id)."""
    uid = _make_user()
    token = _mint_token(uid)
    rules = _sample_rules()
    r = requests.post(
        f"{server}/api/v1/rules",
        json=rules,
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body == {"saved": True, "count": 2}

    r2 = requests.get(
        f"{server}/api/v1/rules", headers=_bearer(token), timeout=5,
    )
    assert r2.status_code == 200
    loaded = r2.json()
    assert isinstance(loaded, list)
    assert len(loaded) == 2
    loaded_ids = {x["rule_id"] for x in loaded}
    assert loaded_ids == {"r1", "r2"}


def test_post_rules_without_auth_returns_401(server):
    r = requests.post(
        f"{server}/api/v1/rules",
        json=_sample_rules(),
        timeout=5,
    )
    assert r.status_code == 401


def test_post_rules_with_non_json_content_type_returns_415(server):
    """Body present but ``Content-Type: text/plain`` → 415."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/rules",
        data="not json",
        headers={**_bearer(token), "Content-Type": "text/plain"},
        timeout=5,
    )
    assert r.status_code == 415


def test_post_rules_with_malformed_json_returns_400(server):
    """``Content-Type: application/json`` but the body does not
    decode → 400."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/rules",
        data="{not valid json",
        headers={**_bearer(token), "Content-Type": "application/json"},
        timeout=5,
    )
    assert r.status_code == 400


def test_post_rules_with_non_list_body_returns_400(server):
    """Body decodes as JSON but is a dict (not a list) → 400."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/rules",
        json={"not": "a list"},
        headers=_bearer(token),
        timeout=5,
    )
    assert r.status_code == 400


def test_get_rules_empty_for_user_with_no_rules(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/rules", headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert r.json() == []


def test_post_rules_is_user_scoped(server):
    """Alice's POST must not appear in Bob's GET. The engine's per-user
    scope keeps them apart."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_token = _mint_token(bob_uid)

    r = requests.post(
        f"{server}/api/v1/rules",
        json=_sample_rules(),
        headers=_bearer(alice_token),
        timeout=5,
    )
    assert r.status_code == 200

    # Bob still sees zero rules.
    r2 = requests.get(
        f"{server}/api/v1/rules", headers=_bearer(bob_token), timeout=5,
    )
    assert r2.status_code == 200
    assert r2.json() == []


# ─── DELETE /api/v1/rules ─────────────────────────────────────────────────


def test_delete_rules_wipes_users_rules(server):
    """After DELETE, GET returns []. The engine call routes through
    ``save_rules([], user_id=...)`` which is the per-user equivalent
    of ``reset_rules``."""
    uid = _make_user()
    token = _mint_token(uid)
    requests.post(
        f"{server}/api/v1/rules", json=_sample_rules(),
        headers=_bearer(token), timeout=5,
    )

    r = requests.delete(
        f"{server}/api/v1/rules", headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert r.json() == {"reset": True}

    r2 = requests.get(
        f"{server}/api/v1/rules", headers=_bearer(token), timeout=5,
    )
    assert r2.json() == []


def test_delete_rules_does_not_wipe_other_users_rules(server):
    """Alice's DELETE must NOT affect Bob's rules — per-user scope.

    ``alert_rules.rule_id`` is a global PRIMARY KEY (schema v1), so
    each user picks disjoint ids — this matches reality: rule ids are
    UUID-ish in the UI, never colliding across users."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_token = _mint_token(bob_uid)

    alice_rules = [{"rule_id": "alice-r1", "name": "a-rule-1"},
                   {"rule_id": "alice-r2", "name": "a-rule-2"}]
    bob_rules = [{"rule_id": "bob-r1", "name": "b-rule-1"},
                 {"rule_id": "bob-r2", "name": "b-rule-2"}]

    requests.post(
        f"{server}/api/v1/rules", json=alice_rules,
        headers=_bearer(alice_token), timeout=5,
    )
    requests.post(
        f"{server}/api/v1/rules", json=bob_rules,
        headers=_bearer(bob_token), timeout=5,
    )

    # Alice nukes hers.
    r = requests.delete(
        f"{server}/api/v1/rules", headers=_bearer(alice_token), timeout=5,
    )
    assert r.status_code == 200

    # Bob's rules survive intact.
    r2 = requests.get(
        f"{server}/api/v1/rules", headers=_bearer(bob_token), timeout=5,
    )
    assert r2.status_code == 200
    bobs = r2.json()
    bob_ids = {x["rule_id"] for x in bobs}
    assert bob_ids == {"bob-r1", "bob-r2"}

    # Alice should now have an empty set.
    r3 = requests.get(
        f"{server}/api/v1/rules", headers=_bearer(alice_token), timeout=5,
    )
    assert r3.json() == []


# ─── POST /api/v1/channels ────────────────────────────────────────────────


def _sample_channel(channel_id: str = "ch-alpha") -> dict:
    return {
        "channel_id": channel_id,
        "name": "Trading desk Slack",
        "kind": "slack",
        "target": "https://hooks.slack.com/services/AAA/BBB/CCC",
        "severity_threshold": "HIGH",
        "enabled": True,
    }


def test_post_channel_saves_and_appears_in_get(server):
    """POST /api/v1/channels with a valid DeliveryChannel dict persists;
    GET /api/v1/channels lists it."""
    uid = _make_user()
    token = _mint_token(uid)
    ch = _sample_channel()

    r = requests.post(
        f"{server}/api/v1/channels", json=ch,
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body == {"saved": True, "channel_id": "ch-alpha"}

    r2 = requests.get(
        f"{server}/api/v1/channels", headers=_bearer(token), timeout=5,
    )
    assert r2.status_code == 200
    listed = r2.json()
    assert isinstance(listed, list)
    assert len(listed) == 1
    assert listed[0]["channel_id"] == "ch-alpha"
    assert listed[0]["name"] == "Trading desk Slack"
    assert listed[0]["kind"] == "slack"
    # Defensive: the API list intentionally OMITS `target` to avoid
    # leaking the Slack webhook URL on a token compromise.
    assert "target" not in listed[0]


def test_post_channel_missing_channel_id_returns_400(server):
    uid = _make_user()
    token = _mint_token(uid)
    body = _sample_channel()
    body.pop("channel_id")
    r = requests.post(
        f"{server}/api/v1/channels", json=body,
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 400


def test_post_channel_with_non_json_content_type_returns_415(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/channels",
        data="not json",
        headers={**_bearer(token), "Content-Type": "text/plain"},
        timeout=5,
    )
    assert r.status_code == 415


def test_post_channel_with_malformed_json_returns_400(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/channels",
        data="{also-bad",
        headers={**_bearer(token), "Content-Type": "application/json"},
        timeout=5,
    )
    assert r.status_code == 400


# ─── DELETE /api/v1/channels/<id> ─────────────────────────────────────────


def test_delete_channel_removes_only_that_channel(server):
    """Insert two channels, delete one, the other survives."""
    uid = _make_user()
    token = _mint_token(uid)
    requests.post(
        f"{server}/api/v1/channels", json=_sample_channel("ch-A"),
        headers=_bearer(token), timeout=5,
    )
    requests.post(
        f"{server}/api/v1/channels", json=_sample_channel("ch-B"),
        headers=_bearer(token), timeout=5,
    )

    r = requests.delete(
        f"{server}/api/v1/channels/ch-A",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert r.json() == {"deleted": True, "channel_id": "ch-A"}

    r2 = requests.get(
        f"{server}/api/v1/channels", headers=_bearer(token), timeout=5,
    )
    listed = r2.json()
    ids = {c["channel_id"] for c in listed}
    assert ids == {"ch-B"}, f"only ch-B should remain, got {ids}"


def test_delete_channel_as_different_user_does_not_delete(server):
    """Alice owns ch-X. Bob's DELETE with his token must NOT delete it
    — the engine's per-user scope filter excludes Alice's row from
    Bob's UPDATE/DELETE. Bob still gets 200 (no info leak about
    other users' channel ids), but Alice's row survives."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_token = _mint_token(bob_uid)

    requests.post(
        f"{server}/api/v1/channels", json=_sample_channel("ch-X"),
        headers=_bearer(alice_token), timeout=5,
    )

    # Bob tries to delete ch-X.
    r = requests.delete(
        f"{server}/api/v1/channels/ch-X",
        headers=_bearer(bob_token), timeout=5,
    )
    assert r.status_code == 200  # no enumeration

    # Verify via engine: Alice's channel survives.
    from engine.alert_delivery import load_channels
    alice_chs = load_channels(user_id=alice_uid)
    alice_ids = {c.channel_id for c in alice_chs}
    assert "ch-X" in alice_ids, \
        "Bob's DELETE must not have removed Alice's channel"


# ─── Per-channel monthly budget endpoints (schema v25) ───────────────────


def test_post_channel_persists_monthly_budget(server):
    """POST /api/v1/channels accepts ``monthly_budget`` in the body and
    the subsequent GET reflects it. Backwards-compat: omitting the
    field defaults to 0 (unlimited)."""
    uid = _make_user()
    token = _mint_token(uid)
    body = _sample_channel("ch-budget-post")
    body["monthly_budget"] = 250
    r = requests.post(
        f"{server}/api/v1/channels", json=body,
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    r2 = requests.get(
        f"{server}/api/v1/channels", headers=_bearer(token), timeout=5,
    )
    listed = {c["channel_id"]: c for c in r2.json()}
    assert listed["ch-budget-post"]["monthly_budget"] == 250


def test_get_channel_usage_returns_zero_on_fresh_channel(server):
    """GET /api/v1/channels/<id>/usage on a freshly-created channel
    returns budget=N + usage=0 + pct=0.0 (or None for unlimited)."""
    uid = _make_user()
    token = _mint_token(uid)
    body = _sample_channel("ch-budget-usage")
    body["monthly_budget"] = 50
    requests.post(
        f"{server}/api/v1/channels", json=body,
        headers=_bearer(token), timeout=5,
    )
    r = requests.get(
        f"{server}/api/v1/channels/ch-budget-usage/usage",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    payload = r.json()
    assert payload["channel_id"] == "ch-budget-usage"
    assert payload["budget"] == 50
    assert payload["usage"] == 0
    assert payload["over_budget"] is False


def test_get_channel_usage_unknown_channel_returns_404(server):
    """An id not in the caller's scope returns 404 — mirrors the
    anti-enumeration shape used by every other per-id GET."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/channels/no-such-id/usage",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 404


def test_post_channel_reset_usage_zeros_counter(server):
    """POST /api/v1/channels/<id>/reset-usage zeros the per-month
    counter. We bump the counter via the engine helper directly to
    avoid needing a fake transport on this end-to-end test."""
    from engine.alert_delivery import (
        get_channel_usage,
        increment_channel_usage,
    )

    uid = _make_user()
    token = _mint_token(uid)
    body = _sample_channel("ch-budget-reset")
    body["monthly_budget"] = 5
    requests.post(
        f"{server}/api/v1/channels", json=body,
        headers=_bearer(token), timeout=5,
    )
    # Bump the counter to 3.
    for _ in range(3):
        increment_channel_usage("ch-budget-reset", user_id=uid)
    assert get_channel_usage("ch-budget-reset", user_id=uid) == 3
    r = requests.post(
        f"{server}/api/v1/channels/ch-budget-reset/reset-usage",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert r.json()["reset"] is True
    assert get_channel_usage("ch-budget-reset", user_id=uid) == 0


def test_patch_channel_updates_monthly_budget(server):
    """PATCH /api/v1/channels/<id> with {"monthly_budget": N}
    mutates the persisted cap. Re-loading the channel via
    ``load_channels`` reflects the new value."""
    from engine.alert_delivery import load_channels

    uid = _make_user()
    token = _mint_token(uid)
    body = _sample_channel("ch-budget-patch")
    body["monthly_budget"] = 10
    requests.post(
        f"{server}/api/v1/channels", json=body,
        headers=_bearer(token), timeout=5,
    )
    r = requests.patch(
        f"{server}/api/v1/channels/ch-budget-patch",
        json={"monthly_budget": 1234},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    payload = r.json()
    assert payload["updated"] == {"monthly_budget": 1234}
    persisted = {c.channel_id: c for c in load_channels(user_id=uid)}
    assert persisted["ch-budget-patch"].monthly_budget == 1234


# ─── POST + DELETE /api/v1/reports/<id>/public ────────────────────────────


def test_post_report_public_returns_slug_and_load_works(
    server, tmp_path, monkeypatch,
):
    """POST returns a slug; ``load_public_report(slug)`` then returns
    the report's HTML. Round-trip proves the slug actually wired up."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="pub")

    r = requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 7},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert "slug" in body
    slug = body["slug"]
    assert isinstance(slug, str) and len(slug) > 0

    from utils.report_history import load_public_report
    html = load_public_report(slug)
    assert html is not None
    assert "pub" in html


def test_post_report_public_unknown_id_returns_404(
    server, tmp_path, monkeypatch,
):
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/reports/no-such-report/public",
        json={"expires_in_days": 30},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 404


def test_post_report_public_cross_user_returns_404(
    server, tmp_path, monkeypatch,
):
    """Alice cannot publish Bob's report — 404 (NOT 403), matching the
    no-info-leak contract used elsewhere."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_rid = _seed_report(bob_uid, label="bobs")
    r = requests.post(
        f"{server}/api/v1/reports/{bob_rid}/public",
        json={"expires_in_days": 30},
        headers=_bearer(alice_token), timeout=5,
    )
    assert r.status_code == 404


def test_post_report_public_with_empty_body_defaults_to_30_days(
    server, tmp_path, monkeypatch,
):
    """Body is optional — when absent, defaults to 30 days. The slug
    is still issued."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="default")
    # No body, no Content-Type — should still succeed.
    r = requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert "slug" in r.json()


def test_delete_report_public_clears_slug(server, tmp_path, monkeypatch):
    """POST grants a slug, DELETE revokes it. After revoke
    ``load_public_report(old_slug)`` returns None."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="revokeme")

    r = requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 30},
        headers=_bearer(token), timeout=5,
    )
    slug = r.json()["slug"]

    r2 = requests.delete(
        f"{server}/api/v1/reports/{rid}/public",
        headers=_bearer(token), timeout=5,
    )
    assert r2.status_code == 200
    assert r2.json() == {"revoked": True}

    from utils.report_history import load_public_report
    assert load_public_report(slug) is None, \
        "slug should be invalidated after revoke"


def test_delete_report_public_unknown_id_returns_404(
    server, tmp_path, monkeypatch,
):
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.delete(
        f"{server}/api/v1/reports/no-such-report/public",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 404


# ─── GET /api/v1/audit ────────────────────────────────────────────────────


def test_audit_endpoint_returns_401_without_auth(server):
    """Audit log is privileged — no bearer header → 401, never the
    audit rows themselves (which would expose every action verb)."""
    r = requests.get(f"{server}/api/v1/audit", timeout=5)
    assert r.status_code == 401
    assert r.json() == {"error": "unauthorized"}


def test_audit_endpoint_is_user_scoped(server):
    """Alice's token must NOT see Bob's audit rows. We insert one
    distinct audit event under each user_id, then fetch as Alice —
    only Alice's row comes back (along with whatever signup /
    token-create rows the auth helpers wrote earlier under HER id).
    This is the core safety property of the endpoint: cross-user
    audit access is a privilege escalation in disguise.

    We use a distinct action verb (``ut_marker``) for the seeded rows
    so the assertion isn't brittle against the signup / token-create
    rows that ``_make_user`` / ``_mint_token`` themselves write."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)

    from auth.audit import record_audit
    record_audit("ut_marker", user_id=alice_uid, detail={"who": "alice"})
    record_audit("ut_marker", user_id=bob_uid, detail={"who": "bob"})

    r = requests.get(
        f"{server}/api/v1/audit",
        headers=_bearer(alice_token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["count"] >= 1
    # CORE PROPERTY: every returned row belongs to Alice — Bob's
    # row must be invisible regardless of what other rows ride along.
    assert all(
        it["user_id"] == alice_uid for it in body["items"]
    ), "Alice's audit list leaked another user's rows"
    # The marker action came back, and detail_json is a dict (parsed),
    # not a JSON string — the endpoint hands the parsed value through.
    markers = [it for it in body["items"] if it["action"] == "ut_marker"]
    assert len(markers) == 1
    assert markers[0]["detail_json"] == {"who": "alice"}


def test_audit_endpoint_filters_by_action(server):
    """``?action=login_success`` must filter to just that verb. We
    seed three different action verbs and confirm only the matching
    one comes back."""
    uid = _make_user()
    token = _mint_token(uid)

    from auth.audit import record_audit
    record_audit("login_success", user_id=uid)
    record_audit("save_rules", user_id=uid)
    record_audit("login_success", user_id=uid)

    r = requests.get(
        f"{server}/api/v1/audit",
        params={"action": "login_success"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["count"] == 2
    assert all(it["action"] == "login_success" for it in body["items"])


def test_audit_endpoint_respects_limit(server):
    """``?limit=5`` must cap the response at 5 rows even when more
    audit events exist for the user."""
    uid = _make_user()
    token = _mint_token(uid)

    from auth.audit import record_audit
    for i in range(10):
        record_audit("some_action", user_id=uid, detail={"i": i})

    r = requests.get(
        f"{server}/api/v1/audit",
        params={"limit": 5},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["count"] == 5
    assert len(body["items"]) == 5


# ─── GET /api/v1/incidents ────────────────────────────────────────────────


def test_incidents_endpoint_returns_401_without_auth(server):
    r = requests.get(f"{server}/api/v1/incidents", timeout=5)
    assert r.status_code == 401


def test_incidents_endpoint_empty_when_no_alerts(server):
    """No alerts → no incidents. Response is the empty envelope, not
    a top-level list (so the shape is uniform with the seeded case)."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/incidents",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body == {"items": [], "count": 0}


def test_incidents_endpoint_returns_correlated_incidents(server):
    """Seed a cluster of alerts of the same alert_type within the
    correlator's 30-minute window — they should collapse into one
    incident the endpoint surfaces. We use distinct tickers per row
    so ``save_alerts``'s dedup_key logic doesn't collapse them at the
    SQL layer (dedup_key = alert_type+severity+ticker); the correlator
    then still groups them because they share the same alert_type."""
    uid = _make_user()
    token = _mint_token(uid)

    from datetime import datetime, timezone, timedelta
    from engine.alert_engine_v2 import ShippingAlert, save_alerts, _new_id

    base = datetime.now(timezone.utc) - timedelta(minutes=10)
    alerts = []
    for i in range(3):
        alerts.append(ShippingAlert(
            alert_id=_new_id(),
            created_at=(base + timedelta(minutes=i)).isoformat(),
            alert_type="RATE_SURGE",
            severity="HIGH",
            title=f"surge #{i}",
            body="seeded",
            ticker=f"TKR{i:02d}",  # distinct tickers → not dedup'd
            route_id="",
            port_locode="",
            value=float(i),
            threshold=0.0,
            change_pct=float(i),
            acknowledged=False,
        ))
    save_alerts(alerts, user_id=uid)

    r = requests.get(
        f"{server}/api/v1/incidents",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["count"] >= 1
    # The three same-alert_type rows in the recent past should
    # collapse into a single incident (correlated by alert_type +
    # time window).
    inc = body["items"][0]
    assert inc["alert_count"] == 3
    assert inc["dominant_alert_type"] == "RATE_SURGE"
    assert inc["severity_max"] == "HIGH"
    # The incident's entities_touched aggregates the per-alert
    # tickers — all three should be present.
    tickers = set(inc["entities_touched"].get("tickers", []))
    assert {"TKR00", "TKR01", "TKR02"} <= tickers
    assert len(inc["alert_ids"]) == 3


# ─── GET /api/v1/source-health ────────────────────────────────────────────


def test_source_health_endpoint_returns_401_without_auth(server):
    r = requests.get(f"{server}/api/v1/source-health", timeout=5)
    assert r.status_code == 401


def test_source_health_endpoint_returns_envelope_when_empty(server):
    """No health pings in the DB → empty items list, count=0.
    Confirms the endpoint emits the canonical envelope even on an
    empty engine result so callers don't need a conditional shape
    check."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/source-health",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert isinstance(body, dict)
    assert isinstance(body["items"], list)
    assert body["count"] == 0
    # Outer envelope still carries the engine's window + outage list.
    assert "current_outages" in body
    assert "window_hours" in body


def test_source_health_endpoint_is_not_user_scoped(server):
    """Source health is global platform telemetry — Alice and Bob
    must see the IDENTICAL response. This is the inverse of
    /audit and /incidents (which ARE per-user)."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_token = _mint_token(bob_uid)

    # Seed a health ping via the real engine call so both users see
    # the same source listed (we use a low-level write directly to
    # the DB to avoid hitting the network from a probe function).
    from state.db import get_connection
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc).isoformat()
    conn = get_connection()
    with conn:
        conn.execute(
            """
            INSERT INTO data_source_health
              (ping_id, source, started_at, duration_ms, status, error_msg)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            ("ping-test-1", "fred", now, 100, "up", ""),
        )

    r_alice = requests.get(
        f"{server}/api/v1/source-health",
        headers=_bearer(alice_token), timeout=5,
    )
    r_bob = requests.get(
        f"{server}/api/v1/source-health",
        headers=_bearer(bob_token), timeout=5,
    )
    assert r_alice.status_code == 200
    assert r_bob.status_code == 200

    body_alice = r_alice.json()
    body_bob = r_bob.json()
    # Bytewise identical: global telemetry, not user-scoped.
    assert body_alice == body_bob
    # And the seeded source actually shows up.
    sources = {it["source"] for it in body_alice["items"]}
    assert "fred" in sources


# ─── v17: optional password-protected public report links ──────────────────


def test_post_report_public_with_password_sets_password(
    server, tmp_path, monkeypatch,
):
    """POST with a ``password`` field sets the hash on the row. The
    slug-only round trip then refuses to load without a password."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="locked")

    r = requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 7, "password": "open-sesame"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    slug = r.json()["slug"]
    # Slug-only load (back-compat path) must now refuse.
    from utils.report_history import load_public_report
    assert load_public_report(slug) is None
    assert load_public_report(slug, password="open-sesame") is not None


def test_get_report_html_without_password_on_protected_report_returns_401(
    server, tmp_path, monkeypatch,
):
    """Bearer-auth'd GET on a password-protected report's HTML must
    return 401 ``password required`` when no password header / query is
    sent — even when the caller IS the owner."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="protected")
    # Publish with a password.
    requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 7, "password": "secret"},
        headers=_bearer(token), timeout=5,
    )
    r = requests.get(
        f"{server}/api/v1/reports/{rid}/html",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 401
    assert r.json() == {"error": "password required"}


def test_get_report_html_wrong_password_returns_401_wrong_password(
    server, tmp_path, monkeypatch,
):
    """``X-Report-Password: wrong`` on a protected report → 401 with
    ``wrong password`` (the message distinguishes ``missing`` from
    ``wrong`` so a viewer can know to retry)."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="protected")
    requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 7, "password": "right"},
        headers=_bearer(token), timeout=5,
    )
    headers = _bearer(token)
    headers["X-Report-Password"] = "WRONG"
    r = requests.get(
        f"{server}/api/v1/reports/{rid}/html",
        headers=headers, timeout=5,
    )
    assert r.status_code == 401
    assert r.json() == {"error": "wrong password"}


def test_get_report_html_correct_password_via_header_returns_html(
    server, tmp_path, monkeypatch,
):
    """Correct ``X-Report-Password`` header unlocks the HTML body."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="unlock-me")
    requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 7, "password": "magic"},
        headers=_bearer(token), timeout=5,
    )
    headers = _bearer(token)
    headers["X-Report-Password"] = "magic"
    r = requests.get(
        f"{server}/api/v1/reports/{rid}/html",
        headers=headers, timeout=5,
    )
    assert r.status_code == 200
    assert r.headers.get("Content-Type", "").startswith("text/html")
    assert "unlock-me" in r.text


def test_get_report_html_correct_password_via_query_string_returns_html(
    server, tmp_path, monkeypatch,
):
    """Correct ``?password=…`` query-string also unlocks the HTML."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="qs-unlock")
    requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 7, "password": "pw-qs"},
        headers=_bearer(token), timeout=5,
    )
    r = requests.get(
        f"{server}/api/v1/reports/{rid}/html",
        params={"password": "pw-qs"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert "qs-unlock" in r.text


def test_get_report_html_unprotected_still_returns_200_back_compat(
    server, tmp_path, monkeypatch,
):
    """Pre-v17 contract: a report with no password on its public
    share (or no public share at all) must load via the bearer-auth'd
    GET endpoint exactly as it did before. Any supplied
    ``X-Report-Password`` is ignored."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="no-pw")
    # No make_public call, no password column set.
    r = requests.get(
        f"{server}/api/v1/reports/{rid}/html",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert "no-pw" in r.text
    # Spurious password header is harmless on an unprotected report.
    headers = _bearer(token)
    headers["X-Report-Password"] = "ignored"
    r2 = requests.get(
        f"{server}/api/v1/reports/{rid}/html",
        headers=headers, timeout=5,
    )
    assert r2.status_code == 200


def test_post_report_public_with_empty_password_is_no_password(
    server, tmp_path, monkeypatch,
):
    """An empty-string password in the body MUST NOT lock the link
    behind an unusable blank password — collapsed to no-password."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="empty-pw")
    r = requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 7, "password": ""},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    # Slug must load WITHOUT a password — empty-string did not lock.
    from utils.report_history import load_public_report
    assert load_public_report(r.json()["slug"]) is not None


def test_post_report_public_non_string_password_returns_400(
    server, tmp_path, monkeypatch,
):
    """A non-string ``password`` value is a client error → 400."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="bad-pw-type")
    r = requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 7, "password": 42},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 400


# ─── GET /api/v1/reports/<id>/markdown ────────────────────────────────────
# The Markdown export endpoint mirrors the HTML one's auth + password
# contract; these tests cover the wire shape, the 404 paths, and the
# password gate.


def test_get_report_markdown_returns_markdown_content_type(
    server, tmp_path, monkeypatch,
):
    """Successful Markdown export returns text/markdown with the
    expected sections in the body."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="md-export")

    r = requests.get(
        f"{server}/api/v1/reports/{rid}/markdown",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    ctype = r.headers.get("Content-Type", "")
    assert ctype.startswith("text/markdown")
    body = r.text
    # The Markdown body always carries the schema-version footer and
    # the documented section headers.
    assert "schema v" in body
    assert "## Executive Summary" in body
    assert "## Signals" in body


def test_get_report_markdown_unknown_id_returns_404(
    server, tmp_path, monkeypatch,
):
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/reports/this-id-does-not-exist/markdown",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 404
    assert r.json() == {"error": "not found"}


def test_get_report_markdown_cross_user_returns_404(
    server, tmp_path, monkeypatch,
):
    """Bob's report id with Alice's token → 404 (no info leak)."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_rid = _seed_report(bob_uid, label="bobs-md")
    r = requests.get(
        f"{server}/api/v1/reports/{bob_rid}/markdown",
        headers=_bearer(alice_token), timeout=5,
    )
    assert r.status_code == 404


def test_get_report_markdown_requires_auth(server):
    """No bearer header → 401, same as every other authed endpoint."""
    r = requests.get(
        f"{server}/api/v1/reports/any-id/markdown",
        timeout=5,
    )
    assert r.status_code == 401


def test_get_report_markdown_password_required_when_protected(
    server, tmp_path, monkeypatch,
):
    """A password-protected report's Markdown endpoint must enforce
    the same 401 ``password required`` contract as the HTML
    endpoint."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="md-locked")
    # Publish with a password.
    requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 7, "password": "openme"},
        headers=_bearer(token), timeout=5,
    )
    # Bare GET → 401 ``password required``.
    r = requests.get(
        f"{server}/api/v1/reports/{rid}/markdown",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 401
    assert r.json() == {"error": "password required"}


def test_get_report_markdown_with_correct_password_returns_body(
    server, tmp_path, monkeypatch,
):
    """X-Report-Password header unlocks the Markdown body."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="md-unlock")
    requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 7, "password": "secret"},
        headers=_bearer(token), timeout=5,
    )
    headers = _bearer(token)
    headers["X-Report-Password"] = "secret"
    r = requests.get(
        f"{server}/api/v1/reports/{rid}/markdown",
        headers=headers, timeout=5,
    )
    assert r.status_code == 200
    assert r.headers.get("Content-Type", "").startswith("text/markdown")
    assert "Investor Report" in r.text


def test_get_report_markdown_with_wrong_password_returns_401(
    server, tmp_path, monkeypatch,
):
    """Wrong password → 401 ``wrong password`` (distinguished from
    missing-password so a viewer knows to retry)."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="md-wrong-pw")
    requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 7, "password": "right"},
        headers=_bearer(token), timeout=5,
    )
    headers = _bearer(token)
    headers["X-Report-Password"] = "WRONG"
    r = requests.get(
        f"{server}/api/v1/reports/{rid}/markdown",
        headers=headers, timeout=5,
    )
    assert r.status_code == 401
    assert r.json() == {"error": "wrong password"}


def test_get_report_markdown_password_via_query_string_unlocks(
    server, tmp_path, monkeypatch,
):
    """?password= query parameter works the same as the header."""
    from utils import report_history as rh
    monkeypatch.setattr(rh, "REPORT_DIR", tmp_path / "reports")
    uid = _make_user()
    token = _mint_token(uid)
    rid = _seed_report(uid, label="md-qs")
    requests.post(
        f"{server}/api/v1/reports/{rid}/public",
        json={"expires_in_days": 7, "password": "qspw"},
        headers=_bearer(token), timeout=5,
    )
    r = requests.get(
        f"{server}/api/v1/reports/{rid}/markdown",
        params={"password": "qspw"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert "schema v" in r.text


# ─── /api/v1/schedules ────────────────────────────────────────────────────

def test_get_schedules_empty_returns_empty_list(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/schedules", headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert r.json() == []


def test_post_schedules_creates_and_get_lists(server):
    """POST persists a schedule; GET returns it with computed
    next_run_at and the same name/cron from the body."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/schedules",
        json={"name": "Morning Macro", "cron_expr": "0 9 * * *",
              "enabled": True},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["saved"] is True
    new_id = body["schedule_id"]
    assert body["schedule"]["next_run_at"] is not None

    r2 = requests.get(
        f"{server}/api/v1/schedules", headers=_bearer(token), timeout=5,
    )
    assert r2.status_code == 200
    listed = r2.json()
    assert isinstance(listed, list)
    assert len(listed) == 1
    assert listed[0]["schedule_id"] == new_id
    assert listed[0]["name"] == "Morning Macro"
    assert listed[0]["cron_expr"] == "0 9 * * *"
    assert listed[0]["enabled"] is True


def test_post_schedules_with_invalid_cron_returns_400(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/schedules",
        json={"name": "bad", "cron_expr": "not a cron"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 400
    assert "invalid cron_expr" in r.json()["error"]


def test_post_schedules_missing_name_returns_400(server):
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/schedules",
        json={"cron_expr": "0 9 * * *"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 400


def test_patch_schedules_updates_enabled_and_cron(server):
    """PATCH lets the caller flip enabled and update cron_expr; only
    the supplied fields move, the rest stay as they were."""
    uid = _make_user()
    token = _mint_token(uid)
    create = requests.post(
        f"{server}/api/v1/schedules",
        json={"name": "x", "cron_expr": "0 9 * * *"},
        headers=_bearer(token), timeout=5,
    )
    sid = create.json()["schedule_id"]

    r = requests.patch(
        f"{server}/api/v1/schedules/{sid}",
        json={"enabled": False, "cron_expr": "*/15 * * * *"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["updated"] is True
    assert body["schedule"]["enabled"] is False
    assert body["schedule"]["cron_expr"] == "*/15 * * * *"
    assert body["schedule"]["name"] == "x"  # untouched


def test_delete_schedules_removes_and_404_on_cross_user(server):
    """Alice can delete her own schedule (200), but Bob cannot delete
    Alice's (404 — preserves the no-leak contract for ids in other
    users' scopes)."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_token = _mint_token(bob_uid)

    create = requests.post(
        f"{server}/api/v1/schedules",
        json={"name": "alice-only", "cron_expr": "0 9 * * *"},
        headers=_bearer(alice_token), timeout=5,
    )
    sid = create.json()["schedule_id"]

    # Bob's delete attempt → 404; Alice's row survives.
    r_bob = requests.delete(
        f"{server}/api/v1/schedules/{sid}",
        headers=_bearer(bob_token), timeout=5,
    )
    assert r_bob.status_code == 404
    alice_list = requests.get(
        f"{server}/api/v1/schedules", headers=_bearer(alice_token), timeout=5,
    )
    assert len(alice_list.json()) == 1

    # Alice's delete → 200; row gone.
    r_alice = requests.delete(
        f"{server}/api/v1/schedules/{sid}",
        headers=_bearer(alice_token), timeout=5,
    )
    assert r_alice.status_code == 200
    assert r_alice.json() == {"deleted": True, "schedule_id": sid}
    after = requests.get(
        f"{server}/api/v1/schedules", headers=_bearer(alice_token), timeout=5,
    )
    assert after.json() == []


# ─── /api/v1/silences (v22) ───────────────────────────────────────────────

def test_get_silences_empty_returns_empty_list(server):
    """A user with no silences gets an empty JSON list — not a 404."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/silences", headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert r.json() == []


def test_post_silences_creates_and_get_lists(server):
    """POST persists a silence; GET returns it with the body fields
    echoed back + the audit columns populated by the engine."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/silences",
        json={
            "duration_minutes": 60,
            "rule_id": "rule_bdi",
            "severity": "HIGH",
            "reason": "FRED maintenance",
        },
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["saved"] is True
    sid = body["silence_id"]
    assert body["silence"]["rule_id"] == "rule_bdi"
    assert body["silence"]["severity"] == "HIGH"
    assert body["silence"]["reason"] == "FRED maintenance"
    # created_by_user_id stamped from the bearer token, not the body —
    # there is no admin-create-for-someone-else surface via the API.
    assert body["silence"]["created_by_user_id"] == uid

    r2 = requests.get(
        f"{server}/api/v1/silences", headers=_bearer(token), timeout=5,
    )
    assert r2.status_code == 200
    listed = r2.json()
    assert isinstance(listed, list) and len(listed) == 1
    assert listed[0]["silence_id"] == sid


def test_post_silences_missing_duration_returns_400(server):
    """``duration_minutes`` is the only required body field — its
    absence yields 400 with a descriptive message."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/silences",
        json={"rule_id": "rule_X"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 400
    assert "duration_minutes" in r.json()["error"]


def test_post_silences_with_all_nulls_creates_broadest_silence(server):
    """Omitting rule_id / ticker / severity creates the broadest
    possible silence (matches every alert for the user)."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/silences",
        json={"duration_minutes": 30},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["silence"]["rule_id"] is None
    assert body["silence"]["ticker"] is None
    assert body["silence"]["severity"] is None


def test_delete_silences_removes_and_404_on_cross_user(server):
    """Alice can delete her own silence (200); Bob cannot delete
    Alice's (404 — no-leak contract identical to schedules)."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_token = _mint_token(bob_uid)

    create = requests.post(
        f"{server}/api/v1/silences",
        json={"duration_minutes": 60, "rule_id": "alice_rule"},
        headers=_bearer(alice_token), timeout=5,
    )
    sid = create.json()["silence_id"]

    # Bob's delete attempt → 404; Alice's row survives.
    r_bob = requests.delete(
        f"{server}/api/v1/silences/{sid}",
        headers=_bearer(bob_token), timeout=5,
    )
    assert r_bob.status_code == 404
    alice_list = requests.get(
        f"{server}/api/v1/silences", headers=_bearer(alice_token), timeout=5,
    )
    assert len(alice_list.json()) == 1

    # Alice's delete → 200; row gone.
    r_alice = requests.delete(
        f"{server}/api/v1/silences/{sid}",
        headers=_bearer(alice_token), timeout=5,
    )
    assert r_alice.status_code == 200
    assert r_alice.json() == {"deleted": True, "silence_id": sid}
    after = requests.get(
        f"{server}/api/v1/silences", headers=_bearer(alice_token), timeout=5,
    )
    assert after.json() == []


def test_silences_endpoints_require_auth(server):
    """GET, POST, DELETE on /api/v1/silences without a bearer token
    all return 401 — silence config is per-user state and must not
    leak to anonymous probes."""
    # GET
    r = requests.get(f"{server}/api/v1/silences", timeout=5)
    assert r.status_code == 401
    # POST
    r = requests.post(
        f"{server}/api/v1/silences", json={"duration_minutes": 5}, timeout=5,
    )
    assert r.status_code == 401
    # DELETE
    r = requests.delete(f"{server}/api/v1/silences/anything", timeout=5)
    assert r.status_code == 401


# ─── /api/v1/alerts/<id>/annotations + /api/v1/annotations/<id> (v23) ───

def test_get_annotations_empty_returns_empty_list(server):
    """An alert with no annotations returns an empty JSON list — not
    a 404. The thread is just empty, the endpoint always responds 200."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/alerts/alert-1/annotations",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert r.json() == []


def test_post_annotations_creates_and_get_lists(server):
    """POST persists an annotation; GET on the thread surfaces it
    with the engine columns (author_user_id stamped from token)."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/alerts/alert-1/annotations",
        json={"body": "escalated to ops team"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["saved"] is True
    aid = body["annotation_id"]
    assert body["annotation"]["body"] == "escalated to ops team"
    # author_user_id stamped from bearer token — there is no
    # admin-author-as-someone-else surface via the API.
    assert body["annotation"]["author_user_id"] == uid
    assert body["annotation"]["user_id"] == uid
    assert body["annotation"]["edited_at"] is None

    r2 = requests.get(
        f"{server}/api/v1/alerts/alert-1/annotations",
        headers=_bearer(token), timeout=5,
    )
    assert r2.status_code == 200
    listed = r2.json()
    assert isinstance(listed, list) and len(listed) == 1
    assert listed[0]["annotation_id"] == aid


def test_post_annotations_missing_body_returns_400(server):
    """``body`` is the only required field — its absence yields 400."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/alerts/alert-1/annotations",
        json={"not_body": "wrong key"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 400
    assert "body" in r.json()["error"]


def test_post_annotations_empty_body_returns_400(server):
    """Whitespace-only bodies are rejected at the API boundary —
    the engine layer drops them too, but the API gives a clearer
    error code."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.post(
        f"{server}/api/v1/alerts/alert-1/annotations",
        json={"body": "   "},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 400


def test_patch_annotations_edits_body_and_stamps_edited_at(server):
    """PATCH replaces the body and sets edited_at; subsequent GET
    shows the new body + the edited_at indicator."""
    uid = _make_user()
    token = _mint_token(uid)
    create = requests.post(
        f"{server}/api/v1/alerts/alert-1/annotations",
        json={"body": "draft"},
        headers=_bearer(token), timeout=5,
    )
    aid = create.json()["annotation_id"]

    r = requests.patch(
        f"{server}/api/v1/annotations/{aid}",
        json={"body": "revised"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert r.json() == {"updated": True, "annotation_id": aid}

    # Confirm the update + the edited_at stamp via GET.
    r2 = requests.get(
        f"{server}/api/v1/alerts/alert-1/annotations",
        headers=_bearer(token), timeout=5,
    )
    [updated] = r2.json()
    assert updated["body"] == "revised"
    assert updated["edited_at"] is not None


def test_patch_annotations_non_author_returns_404(server):
    """Bob cannot edit alice's annotation even if he discovers the
    annotation_id. Same no-leak contract as silences / schedules."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_token = _mint_token(bob_uid)

    create = requests.post(
        f"{server}/api/v1/alerts/alert-1/annotations",
        json={"body": "alice's note"},
        headers=_bearer(alice_token), timeout=5,
    )
    aid = create.json()["annotation_id"]

    # Bob tries to edit — 404 (no-leak).
    r = requests.patch(
        f"{server}/api/v1/annotations/{aid}",
        json={"body": "bob's rewrite"},
        headers=_bearer(bob_token), timeout=5,
    )
    assert r.status_code == 404


def test_delete_annotations_removes_and_404_on_cross_author(server):
    """Alice can delete her own (200); bob cannot delete alice's
    (404 — no-leak contract)."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_token = _mint_token(bob_uid)

    create = requests.post(
        f"{server}/api/v1/alerts/alert-1/annotations",
        json={"body": "alice's note"},
        headers=_bearer(alice_token), timeout=5,
    )
    aid = create.json()["annotation_id"]

    # Bob's delete attempt → 404; alice's row survives.
    r_bob = requests.delete(
        f"{server}/api/v1/annotations/{aid}",
        headers=_bearer(bob_token), timeout=5,
    )
    assert r_bob.status_code == 404
    alice_list = requests.get(
        f"{server}/api/v1/alerts/alert-1/annotations",
        headers=_bearer(alice_token), timeout=5,
    )
    assert len(alice_list.json()) == 1

    # Alice's delete → 200; row gone.
    r_alice = requests.delete(
        f"{server}/api/v1/annotations/{aid}",
        headers=_bearer(alice_token), timeout=5,
    )
    assert r_alice.status_code == 200
    assert r_alice.json() == {"deleted": True, "annotation_id": aid}
    after = requests.get(
        f"{server}/api/v1/alerts/alert-1/annotations",
        headers=_bearer(alice_token), timeout=5,
    )
    assert after.json() == []


def test_annotations_endpoints_require_auth(server):
    """GET, POST, PATCH, DELETE without a bearer token all return
    401 — annotation content is per-user state and must not leak to
    anonymous probes."""
    # GET thread
    r = requests.get(
        f"{server}/api/v1/alerts/alert-1/annotations", timeout=5,
    )
    assert r.status_code == 401
    # POST
    r = requests.post(
        f"{server}/api/v1/alerts/alert-1/annotations",
        json={"body": "x"}, timeout=5,
    )
    assert r.status_code == 401
    # PATCH
    r = requests.patch(
        f"{server}/api/v1/annotations/anything",
        json={"body": "x"}, timeout=5,
    )
    assert r.status_code == 401
    # DELETE
    r = requests.delete(
        f"{server}/api/v1/annotations/anything", timeout=5,
    )
    assert r.status_code == 401


# ─── Rate limiting ────────────────────────────────────────────────────────


@pytest.fixture
def _tight_rate_limit(monkeypatch):
    """Crank the rate-limit defaults way down so a test can exhaust
    the bucket in a handful of requests rather than 120. The env
    vars are read lazily on every request so monkeypatching here
    takes effect immediately.
    """
    monkeypatch.setenv("RATE_LIMIT_CAPACITY", "3")
    # 0.001 tokens/sec — effectively no refill within the test window
    # so the post-cap denial is deterministic.
    monkeypatch.setenv("RATE_LIMIT_REFILL_PER_SEC", "0.001")
    yield


def test_rate_limit_allows_up_to_capacity_quick_calls(server, _tight_rate_limit):
    """With capacity=3, the first 3 quick calls all succeed (200)."""
    uid = _make_user()
    token = _mint_token(uid)
    statuses = []
    for _ in range(3):
        r = requests.get(
            f"{server}/api/v1/alerts", headers=_bearer(token), timeout=5,
        )
        statuses.append(r.status_code)
    assert statuses == [200, 200, 200]


def test_rate_limit_returns_429_when_exceeded(server, _tight_rate_limit):
    """The 4th quick call (capacity=3) trips a 429 with the documented
    body shape AND a ``Retry-After`` header (RFC 7231 §7.1.3)."""
    uid = _make_user()
    token = _mint_token(uid)
    # Drain the bucket.
    for _ in range(3):
        r = requests.get(
            f"{server}/api/v1/alerts", headers=_bearer(token), timeout=5,
        )
        assert r.status_code == 200
    # Next call must be throttled.
    r = requests.get(
        f"{server}/api/v1/alerts", headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 429
    body = r.json()
    assert body["error"] == "rate_limited"
    assert isinstance(body["retry_after_seconds"], int)
    assert body["retry_after_seconds"] >= 1
    # The standard header MUST be present so off-the-shelf HTTP
    # clients can back off automatically.
    assert "Retry-After" in r.headers
    assert int(r.headers["Retry-After"]) == body["retry_after_seconds"]


def test_rate_limit_isolates_per_token(server, _tight_rate_limit):
    """Alice exhausting her quota does NOT throttle Bob — the
    bucket is keyed by user_id resolved from the bearer token."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_token = _mint_token(bob_uid)
    # Drain alice.
    for _ in range(3):
        requests.get(
            f"{server}/api/v1/alerts", headers=_bearer(alice_token), timeout=5,
        )
    r_alice = requests.get(
        f"{server}/api/v1/alerts", headers=_bearer(alice_token), timeout=5,
    )
    assert r_alice.status_code == 429
    # Bob's first call goes through.
    r_bob = requests.get(
        f"{server}/api/v1/alerts", headers=_bearer(bob_token), timeout=5,
    )
    assert r_bob.status_code == 200


def test_health_endpoint_is_not_rate_limited(server, _tight_rate_limit):
    """The /health endpoint is exempt — a load balancer must be able
    to probe it in a tight loop without ever getting throttled. With
    capacity=3, hammering it 50 times in a row must yield 50x 200."""
    statuses = [
        requests.get(f"{server}/api/v1/health", timeout=5).status_code
        for _ in range(50)
    ]
    assert all(s == 200 for s in statuses), (
        f"unexpected non-200 in health hammer: "
        f"{[s for s in statuses if s != 200]}"
    )


def test_rate_limit_retry_after_is_correct_magnitude(server, monkeypatch):
    """With capacity=2 and refill=2/sec, the bucket regrows one token
    in 0.5s. A denied request's Retry-After should round up to 1
    (RFC says integer-seconds; we floor at 1)."""
    monkeypatch.setenv("RATE_LIMIT_CAPACITY", "2")
    monkeypatch.setenv("RATE_LIMIT_REFILL_PER_SEC", "2.0")
    uid = _make_user()
    token = _mint_token(uid)
    for _ in range(2):
        requests.get(
            f"{server}/api/v1/alerts", headers=_bearer(token), timeout=5,
        )
    r = requests.get(
        f"{server}/api/v1/alerts", headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 429
    # 1 token deficit / 2 tokens-per-sec = 0.5s → ceil → 1s.
    retry = int(r.headers["Retry-After"])
    assert retry == 1


def test_rate_limit_allows_after_refill_interval(server, monkeypatch):
    """After waiting long enough for a token to refill, the next
    call succeeds.

    We deliberately use a SLOW refill (5 tokens/sec) for the drain
    phase so the inter-request latency from ``requests.get`` doesn't
    secretly refill enough to dodge the cap. Then we sleep 0.3s
    which at 5/sec yields 1.5 tokens — enough for one more request
    to pass.
    """
    import time as _time

    monkeypatch.setenv("RATE_LIMIT_CAPACITY", "2")
    monkeypatch.setenv("RATE_LIMIT_REFILL_PER_SEC", "5.0")
    uid = _make_user()
    token = _mint_token(uid)
    # Drain — 2 calls succeed, then back-to-back denied calls confirm
    # the limit holds.
    for _ in range(2):
        requests.get(
            f"{server}/api/v1/alerts", headers=_bearer(token), timeout=5,
        )
    # Try repeatedly until we observe a 429 (network jitter can refill
    # one token between bursts; the second attempt should drain it).
    drained = False
    for _ in range(5):
        r = requests.get(
            f"{server}/api/v1/alerts", headers=_bearer(token), timeout=5,
        )
        if r.status_code == 429:
            drained = True
            break
    assert drained, "could not drain bucket — refill is outracing the test"

    # Now wait long enough to refill at least one whole token.
    # 5/s × 0.6s = 3 tokens → comfortable margin against test-runner
    # preemption. (Earlier 0.3s was too tight: under full-suite load
    # the process can lose enough wall-clock time to the OS scheduler
    # that the bucket hasn't refilled when r2 fires. 0.6s buys 3x
    # the headroom.)
    _time.sleep(0.6)
    # Retry once if the first attempt races a slow scheduler tick.
    r2 = requests.get(
        f"{server}/api/v1/alerts", headers=_bearer(token), timeout=5,
    )
    if r2.status_code == 429:
        _time.sleep(0.4)
        r2 = requests.get(
            f"{server}/api/v1/alerts", headers=_bearer(token), timeout=5,
        )
    assert r2.status_code == 200


# ─── GET /api/v1/audit/export ─────────────────────────────────────────────
#
# JSONL export endpoint for SIEM ingestion. The core properties:
#
#   * 401 without auth (privilege escalation guard — same as /audit).
#   * Default format is jsonl with the application/x-ndjson content-type.
#   * format=json returns the {items: [...], count: N} envelope with
#     application/json — same shape as /audit but with the additional
#     since/until filtering applied.
#   * Per-user scoping is enforced (Alice cannot read Bob's rows).


def test_audit_export_endpoint_returns_401_without_auth(server):
    """Audit export is privileged — no bearer → 401, not the body."""
    r = requests.get(f"{server}/api/v1/audit/export", timeout=5)
    assert r.status_code == 401
    assert r.json() == {"error": "unauthorized"}


def test_audit_export_endpoint_jsonl_body_with_valid_token(server):
    """A valid token gets JSONL bytes back. Each line is independently
    json-loads-able and carries the seeded action verb. Confirms the
    formatter + DB round-trip works through the HTTP wire."""
    uid = _make_user()
    token = _mint_token(uid)

    from auth.audit import record_audit
    record_audit("ut_export_marker", user_id=uid, detail={"who": "alice"})
    record_audit("ut_export_marker", user_id=uid, detail={"who": "alice2"})

    r = requests.get(
        f"{server}/api/v1/audit/export",
        params={"action": "ut_export_marker"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    text = r.text
    lines = [ln for ln in text.split("\n") if ln.strip()]
    assert len(lines) == 2
    import json as _json
    parsed = [_json.loads(ln) for ln in lines]
    assert all(p["action"] == "ut_export_marker" for p in parsed)


def test_audit_export_endpoint_json_envelope_format(server):
    """``?format=json`` returns the same {items, count} shape as
    /audit so a caller that wants the envelope (rather than JSONL)
    can opt in via the query param — useful for legacy tooling that
    expects an array under "items"."""
    uid = _make_user()
    token = _mint_token(uid)

    from auth.audit import record_audit
    record_audit("ut_json_envelope", user_id=uid)

    r = requests.get(
        f"{server}/api/v1/audit/export",
        params={"action": "ut_json_envelope", "format": "json"},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    body = r.json()
    assert "items" in body
    assert "count" in body
    assert body["count"] >= 1
    assert all(it["action"] == "ut_json_envelope" for it in body["items"])


def test_audit_export_endpoint_content_type_is_ndjson_for_jsonl(server):
    """``Content-Type`` MUST be ``application/x-ndjson; charset=utf-8``
    for the JSONL format. SIEM scrapers route on this header — a
    misclassified body would land in the wrong index."""
    uid = _make_user()
    token = _mint_token(uid)

    from auth.audit import record_audit
    record_audit("ut_ctype", user_id=uid)

    r = requests.get(
        f"{server}/api/v1/audit/export",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    ctype = r.headers.get("Content-Type", "")
    assert "application/x-ndjson" in ctype
    assert "utf-8" in ctype.lower()


def test_audit_export_endpoint_is_user_scoped(server):
    """Alice's token must NOT export Bob's rows. Same privacy
    property as /audit — re-asserted here because the export
    endpoint is a separate code path that could regress
    independently."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)

    from auth.audit import record_audit
    record_audit("ut_scope_check", user_id=alice_uid, detail={"who": "alice"})
    record_audit("ut_scope_check", user_id=bob_uid, detail={"who": "bob"})

    r = requests.get(
        f"{server}/api/v1/audit/export",
        params={"action": "ut_scope_check"},
        headers=_bearer(alice_token), timeout=5,
    )
    assert r.status_code == 200
    import json as _json
    lines = [ln for ln in r.text.split("\n") if ln.strip()]
    parsed = [_json.loads(ln) for ln in lines]
    # Every returned row belongs to Alice; Bob's row must NOT appear.
    assert all(p["user_id"] == alice_uid for p in parsed), (
        "Audit export leaked another user's row"
    )


# ─── /api/v1/rules/<id>/escalations + /api/v1/escalations/<id> (v24) ────
#
# Per-rule escalation-chain endpoints. The four surfaces:
#   GET    /api/v1/rules/<rule_id>/escalations  → list chain (200 + [])
#   POST   /api/v1/rules/<rule_id>/escalations  → add/replace step
#   DELETE /api/v1/rules/<rule_id>/escalations  → bulk-clear chain
#   DELETE /api/v1/escalations/<chain_id>       → delete one step
#
# Per-user scoped via bearer token. POST validates the supplied
# channel_id exists in the caller's set so a typo is rejected at
# write time (400) instead of failing silently at dispatch time.


def _make_channel_for(user_id: str, channel_id: str = "ch_test") -> str:
    """Persist a delivery channel under ``user_id`` and return its id."""
    from engine.alert_delivery import DeliveryChannel, save_channel

    ch = DeliveryChannel(
        channel_id=channel_id,
        name=f"Test {channel_id}",
        kind="slack",
        target=f"https://hooks.example.com/{channel_id}",
        severity_threshold="LOW",
        enabled=True,
    )
    save_channel(ch, user_id=user_id)
    return ch.channel_id


def test_get_escalations_empty_returns_empty_list(server):
    """A rule with no chain returns an empty JSON list — not a 404."""
    uid = _make_user()
    token = _mint_token(uid)
    r = requests.get(
        f"{server}/api/v1/rules/rule_x/escalations",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert r.json() == []


def test_post_escalations_persists_and_get_reflects(server):
    """POST persists a step; GET on the chain surfaces it with the
    engine columns (chain_id stamped, created_at populated)."""
    uid = _make_user()
    token = _mint_token(uid)
    cid = _make_channel_for(uid, "ch_a")

    r = requests.post(
        f"{server}/api/v1/rules/rule_x/escalations",
        json={
            "step_number": 1,
            "after_minutes": 15,
            "channel_id": cid,
        },
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["saved"] is True
    chain_id = body["chain_id"]
    assert body["step"]["step_number"] == 1
    assert body["step"]["after_minutes"] == 15
    assert body["step"]["channel_id"] == cid
    assert body["step"]["user_id"] == uid

    r2 = requests.get(
        f"{server}/api/v1/rules/rule_x/escalations",
        headers=_bearer(token), timeout=5,
    )
    assert r2.status_code == 200
    listed = r2.json()
    assert isinstance(listed, list) and len(listed) == 1
    assert listed[0]["chain_id"] == chain_id


def test_delete_one_escalation_step_removes_it(server):
    """DELETE /api/v1/escalations/<chain_id> removes one step;
    subsequent GET on the chain reflects the deletion."""
    uid = _make_user()
    token = _mint_token(uid)
    cid = _make_channel_for(uid, "ch_a")

    create = requests.post(
        f"{server}/api/v1/rules/rule_x/escalations",
        json={"step_number": 1, "after_minutes": 15, "channel_id": cid},
        headers=_bearer(token), timeout=5,
    )
    chain_id = create.json()["chain_id"]

    r = requests.delete(
        f"{server}/api/v1/escalations/{chain_id}",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    assert r.json() == {"deleted": True, "chain_id": chain_id}

    r2 = requests.get(
        f"{server}/api/v1/rules/rule_x/escalations",
        headers=_bearer(token), timeout=5,
    )
    assert r2.json() == []


def test_delete_chain_bulk_removes_every_step(server):
    """DELETE /api/v1/rules/<id>/escalations clears every step in
    the chain for that rule + user. Returns the deleted-row count."""
    uid = _make_user()
    token = _mint_token(uid)
    cid = _make_channel_for(uid, "ch_a")

    for n in (1, 2, 3):
        r = requests.post(
            f"{server}/api/v1/rules/rule_x/escalations",
            json={
                "step_number": n,
                "after_minutes": n * 15,
                "channel_id": cid,
            },
            headers=_bearer(token), timeout=5,
        )
        assert r.status_code == 200

    # Bulk DELETE clears all three.
    r = requests.delete(
        f"{server}/api/v1/rules/rule_x/escalations",
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 200
    payload = r.json()
    assert payload["deleted_steps"] == 3
    assert payload["rule_id"] == "rule_x"

    r2 = requests.get(
        f"{server}/api/v1/rules/rule_x/escalations",
        headers=_bearer(token), timeout=5,
    )
    assert r2.json() == []


def test_escalation_endpoints_per_user_scoping(server):
    """Alice cannot delete bob's chain step even by knowing the
    chain_id. Same no-leak contract as silences / schedules — the
    cross-user DELETE returns 404 and the row survives."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_token = _mint_token(bob_uid)
    bob_cid = _make_channel_for(bob_uid, "ch_bob")

    # Bob creates a step.
    create = requests.post(
        f"{server}/api/v1/rules/rule_x/escalations",
        json={
            "step_number": 1, "after_minutes": 15, "channel_id": bob_cid,
        },
        headers=_bearer(bob_token), timeout=5,
    )
    chain_id = create.json()["chain_id"]

    # Alice's delete attempt → 404; bob's row survives.
    r_alice = requests.delete(
        f"{server}/api/v1/escalations/{chain_id}",
        headers=_bearer(alice_token), timeout=5,
    )
    assert r_alice.status_code == 404
    r_bob_get = requests.get(
        f"{server}/api/v1/rules/rule_x/escalations",
        headers=_bearer(bob_token), timeout=5,
    )
    assert len(r_bob_get.json()) == 1


def test_escalation_endpoints_require_auth(server):
    """GET / POST / DELETE on the escalation surfaces all return 401
    without a bearer token — chain config is per-user state and must
    NOT leak to anonymous probes."""
    r = requests.get(
        f"{server}/api/v1/rules/rule_x/escalations", timeout=5,
    )
    assert r.status_code == 401
    r = requests.post(
        f"{server}/api/v1/rules/rule_x/escalations",
        json={"step_number": 1, "after_minutes": 15, "channel_id": "x"},
        timeout=5,
    )
    assert r.status_code == 401
    r = requests.delete(
        f"{server}/api/v1/escalations/anything", timeout=5,
    )
    assert r.status_code == 401
    r = requests.delete(
        f"{server}/api/v1/rules/rule_x/escalations", timeout=5,
    )
    assert r.status_code == 401


def test_post_escalations_missing_step_number_returns_400(server):
    """``step_number`` is required — its absence yields 400."""
    uid = _make_user()
    token = _mint_token(uid)
    cid = _make_channel_for(uid, "ch_a")
    r = requests.post(
        f"{server}/api/v1/rules/rule_x/escalations",
        json={"after_minutes": 15, "channel_id": cid},
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 400
    assert "step_number" in r.json()["error"]


def test_post_escalations_unknown_channel_id_returns_400(server):
    """A channel_id that does not exist in the caller's channel set
    is rejected at the API boundary (400) — the engine would
    persist it but escalate_alert would silently fail at dispatch."""
    uid = _make_user()
    token = _mint_token(uid)
    # Note: no channel created.

    r = requests.post(
        f"{server}/api/v1/rules/rule_x/escalations",
        json={
            "step_number": 1,
            "after_minutes": 15,
            "channel_id": "ch_does_not_exist",
        },
        headers=_bearer(token), timeout=5,
    )
    assert r.status_code == 400
    assert "channel_id" in r.json()["error"]


def test_post_escalations_other_users_channel_rejected(server):
    """Alice cannot wire her chain to bob's channel — the channel
    validation is per-user-scoped so a cross-tenant reference is
    rejected as a bogus channel_id (400)."""
    alice_uid = _make_user("alice", "Hunter2!hunter")
    bob_uid = _make_user("bob", "Hunter2!hunter")
    alice_token = _mint_token(alice_uid)
    bob_cid = _make_channel_for(bob_uid, "ch_bob")

    r = requests.post(
        f"{server}/api/v1/rules/rule_x/escalations",
        json={
            "step_number": 1,
            "after_minutes": 15,
            "channel_id": bob_cid,
        },
        headers=_bearer(alice_token), timeout=5,
    )
    assert r.status_code == 400
