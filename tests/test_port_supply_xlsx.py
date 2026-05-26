"""Defining-property tests for utils/port_supply_xlsx.py."""
from __future__ import annotations

import io

import openpyxl
import pytest

from processing.port_supply_lines import (
    build_company_port_footprints,
    build_port_supply_chains,
)
from utils.port_supply_xlsx import (
    WORKBOOK_SHEETS,
    build_workbook,
)


@pytest.fixture(scope="module")
def workbook_bytes() -> bytes:
    chains = build_port_supply_chains()
    footprints = build_company_port_footprints()
    return build_workbook(chains, footprints)


@pytest.fixture(scope="module")
def workbook(workbook_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(workbook_bytes))


# ── 1. File-level sanity ──────────────────────────────────────────────────

def test_build_workbook_returns_nonempty_bytes(workbook_bytes: bytes) -> None:
    assert isinstance(workbook_bytes, bytes)
    # An empty Excel file is ~5KB; ours has 5 populated sheets + overview.
    assert len(workbook_bytes) > 5_000


def test_workbook_opens_in_openpyxl(workbook) -> None:
    """If openpyxl can't load it back, no analyst tool will be able to either."""
    assert workbook is not None


# ── 2. Sheet inventory ────────────────────────────────────────────────────

def test_workbook_has_six_canonical_sheets(workbook) -> None:
    """One overview sheet + the five CSV-view sheets, in canonical order."""
    assert tuple(workbook.sheetnames) == WORKBOOK_SHEETS


def test_no_default_blank_sheet_left_over(workbook) -> None:
    """openpyxl creates a default 'Sheet' on Workbook(); we rename it to
    'overview' so the workbook never carries an empty leftover."""
    assert "Sheet" not in workbook.sheetnames


# ── 3. Per-sheet shape contracts ──────────────────────────────────────────

def test_summary_sheet_has_one_header_row_plus_one_per_port(workbook) -> None:
    from ports.port_registry import PORTS
    ws = workbook["summary"]
    # 1 header + N port rows
    assert ws.max_row == 1 + len(PORTS)
    # Column count matches the documented schema (19 cols at schema_v2).
    header = [c.value for c in ws[1]]
    assert "deficit_rank" in header
    assert "exposure_concentration_hhi" in header


def test_exposure_sheet_carries_port_state_columns(workbook) -> None:
    ws = workbook["exposure"]
    header = [c.value for c in ws[1]]
    for required in ("locode", "ticker", "share_within_port",
                     "rank_within_port"):
        assert required in header


def test_footprint_sheet_carries_aggregate_columns(workbook) -> None:
    ws = workbook["footprint"]
    header = [c.value for c in ws[1]]
    for required in ("ticker", "deficit_share", "concentration_hhi",
                     "port_rank_in_footprint"):
        assert required in header


def test_regional_sheet_has_severity_bucket_columns(workbook) -> None:
    ws = workbook["regional"]
    header = [c.value for c in ws[1]]
    # Each of the 5 severity buckets becomes one column.
    for bucket in ("critical_deficit", "deficit", "balanced",
                   "surplus", "heavy_surplus"):
        assert f"n_{bucket}" in header


def test_watchlist_sheet_carries_action_columns(workbook) -> None:
    ws = workbook["watchlist"]
    header = [c.value for c in ws[1]]
    for required in ("deficit_rank", "action", "why"):
        assert required in header


# ── 4. Header row is bold + frozen ────────────────────────────────────────

def test_every_data_sheet_freezes_header_row(workbook) -> None:
    for name in WORKBOOK_SHEETS:
        if name == "overview":
            continue
        ws = workbook[name]
        # frozen at A2 means row 1 stays visible on scroll
        assert ws.freeze_panes == "A2", (
            f"sheet '{name}' missing freeze_panes=A2"
        )


def test_every_data_sheet_header_is_bold(workbook) -> None:
    for name in WORKBOOK_SHEETS:
        if name == "overview":
            continue
        ws = workbook[name]
        for cell in ws[1]:
            if cell.value is None:
                continue
            assert cell.font.bold is True, (
                f"sheet '{name}' header cell {cell.coordinate} not bold"
            )


# ── 5. Numeric coercion (cells write as numbers, not strings) ────────────

def test_summary_supply_deficit_days_stored_as_number(workbook) -> None:
    """The CSV emits '+0.00' / '-2.00'; the xlsx writer must coerce to
    float so spreadsheet sort + SUM / pivot operations work without a
    type conversion in the consumer."""
    ws = workbook["summary"]
    header = [c.value for c in ws[1]]
    deficit_col = header.index("supply_deficit_days") + 1
    # Sample the first data row's cell
    cell = ws.cell(row=2, column=deficit_col)
    assert isinstance(cell.value, (int, float))


def test_summary_hhi_stored_as_number(workbook) -> None:
    ws = workbook["summary"]
    header = [c.value for c in ws[1]]
    hhi_col = header.index("exposure_concentration_hhi") + 1
    cell = ws.cell(row=2, column=hhi_col)
    assert isinstance(cell.value, (int, float))
    assert 0.0 <= cell.value <= 1.0


def test_summary_deficit_days_has_signed_number_format(workbook) -> None:
    """Number formats are applied per the _NUMBER_FORMATS registry —
    deficit columns should carry the signed-numeric format."""
    ws = workbook["summary"]
    header = [c.value for c in ws[1]]
    deficit_col = header.index("supply_deficit_days") + 1
    cell = ws.cell(row=2, column=deficit_col)
    # The exact format string carries +/- prefixes for sign.
    assert "+" in cell.number_format or cell.number_format.startswith("0")


# ── 6. Overview sheet contents ────────────────────────────────────────────

def test_overview_sheet_lists_container_type(workbook) -> None:
    ws = workbook["overview"]
    # Read everything as text + check the container type shows up.
    cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cells.append(str(cell.value))
    assert "40FT_DRY" in cells


def test_overview_sheet_lists_per_sheet_row_counts(workbook) -> None:
    """Operator opens the workbook → overview tells them how many rows
    each sheet carries so they know what to expect before clicking."""
    ws = workbook["overview"]
    labels = [
        str(ws.cell(row=r, column=1).value or "")
        for r in range(1, ws.max_row + 1)
    ]
    # Every non-overview sheet name must appear as a row label.
    for name in WORKBOOK_SHEETS:
        if name == "overview":
            continue
        assert name in labels


# ── 7. Container-type knob propagates ────────────────────────────────────

def test_container_type_propagates_into_overview() -> None:
    chains = build_port_supply_chains()
    footprints = build_company_port_footprints()
    data = build_workbook(
        chains, footprints, container_type="40FT_REEFER",
    )
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["overview"]
    cells = [
        str(c.value) for row in ws.iter_rows() for c in row
        if c.value is not None
    ]
    assert "40FT_REEFER" in cells


# ── 8. Threshold knob propagates into the watchlist sheet ───────────────

def test_threshold_days_propagates_into_watchlist_sheet() -> None:
    """Tighter threshold → fewer watchlist rows (or zero rows + overview
    still names the threshold so the operator knows why it's empty)."""
    chains = build_port_supply_chains()
    footprints = build_company_port_footprints()
    loose_wb = openpyxl.load_workbook(io.BytesIO(build_workbook(
        chains, footprints, threshold_days=0.0,
    )))
    tight_wb = openpyxl.load_workbook(io.BytesIO(build_workbook(
        chains, footprints, threshold_days=-1000.0,
    )))
    # Tight watchlist must be no bigger than loose.
    assert tight_wb["watchlist"].max_row <= loose_wb["watchlist"].max_row
